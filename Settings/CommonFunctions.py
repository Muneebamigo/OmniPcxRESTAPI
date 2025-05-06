'''
Created on April 07, 2025

@author: Muneeb.ahmad
'''

import time
import random
import string
import os, pypyodbc
from datetime import datetime
import platform
import string
from random import *
import random
from openpyxl import load_workbook
import openpyxl.styles.alignment as XLStyle

from Settings import dataFunction as DataFunction
from InputDataFiles import SeleniumConfigration as SC
from InputDataFiles import InputData
from openpyxl.styles import Alignment, Font, PatternFill

# from test_Provisioning import test_60_purpleLicesne_auth as Auth


#from autopylogger import init_logging

# ds = DataFunction.DataStorage()
sc = SC.SeleniumConfig()
ssinputdata = InputData.InputData()
class CommonFunctions():
    

    
    starttime = time.process_time()
    PrereqTestCasesStatusUpdate = False # True or False
    SuccessMessage = "Test Case Passed Successfully"
    FailedMessage = "Failed"
    ExecutionDate = str(datetime.now().date())
    ExecutionTime = str(time.strftime("%H:%M:%S", time.localtime()))
    platformsystem = str(platform.system())
    platformrelease = str(platform.release())
    WindowServer = str(platformsystem+platformrelease)
    SystemUser = os.getlogin()
    #Domain = 'http://172.20.11.253/opcxrrestapi'
    Domain = ssinputdata.Domain
    # purple_domain =ssinputdata.purple_domain
    
    
    #authkey_server = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJBbWlnb1NvZnR3YXJlIiwic3ViIjoiT1BDWFJBUEkiLCJlbWFpbCI6ImFzc2lzdGFuY2VAYW1pZ28tc29mdHdhcmUuY29tIiwicm9sZSI6IkludGVncmF0b3IiLCJodHRwOi8vc2NoZW1hcy5taWNyb3NvZnQuY29tL3dzLzIwMDgvMDYvaWRlbnRpdHkvY2xhaW1zL2lzcGVyc2lzdGVudCI6IlRydWUiLCJpYXQiOjE2MzQzMDAyMzUsImh0dHA6Ly9zY2hlbWFzLm1pY3Jvc29mdC5jb20vd3MvMjAwOC8wNi9pZGVudGl0eS9jbGFpbXMvdmVyc2lvbiI6IlByb2R1Y3Rpb24iLCJleHAiOjE2MzQzMDIwMzUsImF1ZCI6Ik9QQ1hSQVBJIiwiTElJIjoiVHJ1ZSIsIlVOU1QiOiJhZG1pbiIsIlVUWVBFIjoiU2VydmVyIiwiU0lURUNPREUiOiJDb25maWciLCJMSURPUyI6IjAwNDI5LTAwMDAwLTAwMDAxLUFBODIwIiwiTVBUIjoiRmFsc2UiLCJleHBpcmVzX2F0IjoiMTYzNDMwMjAzNSIsIlJUQiI6IkZhbHNlIiwiZXhwaXJlc19taW51dGVzIjoiMzAiLCJuYmYiOjE2MzQzMDAyMzV9.KdtWOq2LzoPadZ3qRdCQQ_yEN1DK7Q8EJWj-4bZ5gE4'
    #authkey_Site = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJBbWlnb1NvZnR3YXJlIiwic3ViIjoiT1BDWFJBUEkiLCJlbWFpbCI6ImFzc2lzdGFuY2VAYW1pZ28tc29mdHdhcmUuY29tIiwicm9sZSI6IkludGVncmF0b3IiLCJodHRwOi8vc2NoZW1hcy5taWNyb3NvZnQuY29tL3dzLzIwMDgvMDYvaWRlbnRpdHkvY2xhaW1zL2lzcGVyc2lzdGVudCI6IlRydWUiLCJpYXQiOjE2MzQzMDAyMzYsImh0dHA6Ly9zY2hlbWFzLm1pY3Jvc29mdC5jb20vd3MvMjAwOC8wNi9pZGVudGl0eS9jbGFpbXMvdmVyc2lvbiI6IlByb2R1Y3Rpb24iLCJleHAiOjE2MzQzMDIwMzYsImF1ZCI6Ik9QQ1hSQVBJIiwiTElJIjoiVHJ1ZSIsIlVOU1QiOiJhZG1pbiIsIlVUWVBFIjoiU2l0ZSIsIlNJVEVDT0RFIjoiMDEwMDAxIiwiTElET1MiOiIwMDQyOS0wMDAwMC0wMDAwMS1BQTgyMCIsIk1QVCI6IkZhbHNlIiwiZXhwaXJlc19hdCI6IjE2MzQzMDIwMzYiLCJSVEIiOiJGYWxzZSIsImV4cGlyZXNfbWludXRlcyI6IjMwIiwibmJmIjoxNjM0MzAwMjM2fQ.AwWrNM_Yv4TAHfM86-6-q5bTBEJL_NsEMbpxxIKYs1s'
    #authkey_Site_UserToken = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJBbWlnb1NvZnR3YXJlIiwic3ViIjoiT1BDWFJBUEkiLCJlbWFpbCI6ImFzc2lzdGFuY2VAYW1pZ28tc29mdHdhcmUuY29tIiwicm9sZSI6IlVzZXIiLCJodHRwOi8vc2NoZW1hcy5taWNyb3NvZnQuY29tL3dzLzIwMDgvMDYvaWRlbnRpdHkvY2xhaW1zL2lzcGVyc2lzdGVudCI6IlRydWUiLCJpYXQiOjE2MzQzMDAyNDAsImh0dHA6Ly9zY2hlbWFzLm1pY3Jvc29mdC5jb20vd3MvMjAwOC8wNi9pZGVudGl0eS9jbGFpbXMvdmVyc2lvbiI6IlByb2R1Y3Rpb24iLCJleHAiOjE2MzQzMDIwNDAsImF1ZCI6Ik9QQ1hSQVBJIiwiTElJIjoiVHJ1ZSIsIlVOU1QiOiJhZG1pbiIsIlVUWVBFIjoiU2l0ZSIsIlNJVEVDT0RFIjoiMDEwMDAxIiwiTElET1MiOiIwMDQyOS0wMDAwMC0wMDAwMS1BQTgyMCIsIk1QVCI6IkZhbHNlIiwiZXhwaXJlc19hdCI6IjE2MzQzMDIwNDAiLCJSVEIiOiJGYWxzZSIsImV4cGlyZXNfbWludXRlcyI6IjMwIiwibmJmIjoxNjM0MzAwMjQwfQ.SYxDSmdXBRnJdngC0U2B-wHzWmm6SCBnScxPizSAK4U'
    
    
    #authuser = 'admin'
    # authuser = ssinputdata.authuser
    # purple_authuser= ssinputdata.purple_user
    # purple_masterToken = ssinputdata.purple_masterToken
    
    # #authpassword='1234567a'
    # authpassword=ssinputdata.authpassword
    # purple_authpassword = ssinputdata.purple_ath_pwd
    # Output Result File path
    #OutPutFilePath = 'C:\\Users\\nouman.ijaz\\Documents\\NumanIjaz\\workspace\\R2.4.0.11-OmniPCX RECORD REST API Automated Tests Sheet.xlsx'

    # OutPutFilePath = os.path.dirname(os.path.abspath(__file__)).replace("RestAPI", "")+ssinputdata.ExcelSheetName
    curr_working_directory = os.getcwd()
    file_path = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
    OutPutFilePath= os.path.join(file_path ,ssinputdata.ExcelSheetName)
    # OutPutFilePath = os.path.join(r"C:\Users\Administrator\PycharmProjects\Automation" ,ssinputdata.ExcelSheetName)




    #mylogger = init_logging(log_name="myfirstlogger", log_directory="logsdir")

    def authkey_server(self):
        return sc.get_autoken()[0]
        
    def  authkey_Site(self):
        return sc.get_autoken()[1]
        
    def authkey_Site_UserToken(self):
        return sc.get_autoken()[2]
    
    
    def Header(self,module,testcase,description):
        
        #Print Header__name__
        print('------------------------------Starting------------------------------')
        print('------------------------------' + module + '------------------------------')
        print('------------------------------' + testcase + '------------------------------' )
        print('------------------------------' + description + '------------------------------')
    
    # Tenant DB Connectivity
    def DBConnectivity(self):
        connection = pypyodbc.connect(
            'Driver={PostgreSQL Unicode};'
            'Server=172.20.13.30;'
            'Database=opcxr_tenant_010001_apri_253;'
            'uid=postgres;pwd=0mniPcx'
        )
        cursor = connection.cursor()

        return cursor
    
    # Config DB Connectivity
    def StringDBConnectivity(self):
        connection = pypyodbc.connect(
            'Driver={PostgreSQL Unicode};'
            'Server=172.20.13.30;'
            'Database=opcxr_config_apri_253;'
            'uid=postgres;pwd=0mniPcx'
        )
        cursor = connection.cursor()
        return cursor

    def GenrateValidIPString(self):
        
        '''Generate only integers Or IP Address'''
        first = str(randrange(100,255))
        second = str(randint(1 , 255))
        third = str(randint(1 , 255))
        fourth = str(randint(1 , 255))
        ValidIP=str(''+first+'.'+second+'.'+third+'.'+fourth+'')
        
        return ValidIP
    
    def GenrateSimpleStringLimit10(self):
        
        Simplestring = "".join([random.choice(string.ascii_uppercase) for x in range(6)])
        SimpleString=''+Simplestring+'test'
        
        return SimpleString

    # def GenrateValidPasswordString(self):
    #
    #     upper = "".join([random.choice(string.ascii_uppercase) for x in range(3)])
    #     lower = "".join([random.choice(string.ascii_lowercase) for x in range(3)])
    #     numeric = str(randrange(10,99))
    #     ValidPassword = upper+lower+numeric
    #
    #     return ValidPassword
    # updated method password genrate
    def GenrateValidPasswordString(self):
        # Uppercase letters (3 characters)
        upper = "".join([random.choice(string.ascii_uppercase) for _ in range(3)])

        # Lowercase letters (3 characters)
        lower = "".join([random.choice(string.ascii_lowercase) for _ in range(3)])

        # Numeric characters (2 digits)
        numeric = str(random.randint(10, 99))  # Random number between 10 and 99

        # Special characters (1 character)
        special_characters = "!@#$%^&*()-_=+[]{}|;:,.<>?/`~"  # Add your desired special characters here
        special = random.choice(special_characters)

        # Combine all parts
        valid_password = upper + lower + numeric + special

        # Shuffle the characters to ensure random order
        valid_password = ''.join(random.sample(valid_password, len(valid_password)))

        return valid_password

    def GenerateValidExtension(self):
        
        return  str(random.randint(1000 , 99999  ))
    
    def GenrateValidMac(self):
        first = str(randrange(10 , 90))
        second = str(randint(10 , 90))
        third = str(randint(10 , 90))
        fourth = str(randint(10 , 90))
        fifth = str(randint(10 , 90))
        sixth = str(randint(10 , 90))
        
        ValidMac =str(''+first+':'+second+':'+third+':'+fourth+':'+fifth+':'+sixth+'') 
        return ValidMac
    
    def GenrateDesc250(self):
        
        GenrateDesc250 = "".join([random.choice(string.ascii_uppercase) for x in range(255)])
        return GenrateDesc250
        
    def GenerateSpecialChar(self):
        
        SpecialChar  = ''.join([random.choice(string.ascii_letters + string.digits + string.punctuation ) for n in range(10)])
        return SpecialChar
    
    def GenerateEmail(self):
        
        username = "".join([random.choice(string.ascii_lowercase) for x in range(8)])
        domain = "".join([random.choice(string.ascii_lowercase) for x in range(5)])
        
        Email = ''+username+'@'+domain+'.com'
        
        return Email

    # def get_purple_auth(self):
    #     auth = Auth.PurpleLicenseAuth()
    #     get_token = auth.testcase_01_get_session_token(True)
    #     return get_token

    def parse_application_id(self,data):
        data_response = data.split(':')
        project_id = data_response[2].split(',')[0]
        project_id = project_id.replace('"', "")
        application_id = data_response[3].split('"')[1]
        return project_id, application_id

    # this function is updated for add values in the sheet through script //muneeb.ahmad
    def UpdateExcelTestCase(self, Sheetname, TestCaseID, URL, Parameters, status, starttime, resp,
                            TestDescription="", Methods="", Steps="", ExpectedResult="",
                            ExpectedProcessingTime=0,
                            ExpectedResponseJSON=None, ExpectedCode=None):

        showcode = str(resp['ResponseCode'])

        wb = load_workbook('' + CommonFunctions.OutPutFilePath + '')
        ws = wb[Sheetname]

        first_column = ws['B']
        del Parameters["AuthToken"]

        for x in range(len(first_column)):
            if (first_column[x].value) == TestCaseID:
                row = x + 1

                # New fields (fill missing columns)
                # Apply alignment and style to new fields
                center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

                ws.cell(row=row, column=3).alignment = center_wrap
                ws.cell(row=row, column=3).value = TestDescription

                ws.cell(row=row, column=4).alignment = center_wrap
                ws.cell(row=row, column=4).font = Font(bold=True)
                ws.cell(row=row, column=4).value = Methods

                ws.cell(row=row, column=5).alignment = center_wrap
                ws.cell(row=row, column=5).value = Steps

                ws.cell(row=row, column=6).alignment = center_wrap
                ws.cell(row=row, column=6).value = ExpectedResult

                ws.cell(row=row, column=10).alignment = center_wrap
                ws.cell(row=row, column=10).value = ExpectedProcessingTime

                ws.cell(row=row, column=15).alignment = center_wrap
                ws.cell(row=row, column=15).value = str(ExpectedResponseJSON)

                ws.cell(row=row, column=17).alignment = center_wrap
                ws.cell(row=row, column=17).value = ExpectedCode

                # Already existing functionality
                ws.cell(row=row, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                         wrap_text=True)
                ws.cell(row=row, column=7).value = 'Header \n' + str(Parameters).replace(',', '\n')

                ws.cell(row=row, column=8).value = CommonFunctions.ExecutionDate
                ws.cell(row=row, column=9).value = str(time.strftime("%H:%M:%S", time.localtime()))

                ProcessingTime = float(str((time.process_time() - starttime + 2)))
                ws.cell(row=row, column=11).value = ProcessingTime

                ws.cell(row=row, column=13).value = CommonFunctions.SystemUser
                ws.cell(row=row, column=14).value = CommonFunctions.WindowServer

                ws.cell(row=row, column=16).value = str(resp)

                ws.cell(row=row, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                ws.cell(row=row, column=18).value = showcode

                if status == 'Passed':
                    ws.cell(row=row, column=19).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                   fill_type='solid')
                else:
                    ws.cell(row=row, column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                   fill_type='solid')

                ws.cell(row=row, column=19).value = status
                break

        wb.save('' + CommonFunctions.OutPutFilePath + '')

        # Console Print
        print("-------------------Test Results------------------")
        print("Test Case ID : " + TestCaseID)
        print("URL  : " + URL)
        print("Header  : ")
        print(Parameters)
        print("Test Status  : " + status)
        print("Response  : ")
        print(resp)
        print("--------------------------------------------")

    def UpdateExcelTestCasepurple(self, Sheetname, TestCaseID, URL, Parameters, status, starttime, resp):

        showcode = str(resp['responseCode'])

        wb = load_workbook('' + CommonFunctions.OutPutFilePath + '')
        wb.sheetnames
        ws = wb[Sheetname]

        first_column = ws['B']
        del Parameters["AuthToken"]
        for x in range(len(first_column)):
            if (first_column[x].value) == TestCaseID:

                ws.cell(row=x + 1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True, wrapText=True)
                ws.cell(row=x + 1, column=7).value = 'Header \n' + str(Parameters).replace(',', '\n')
                # ws.cell(row=x + 1, column=5).value = URL
                ws.cell(row=x + 1, column=8).value = CommonFunctions.ExecutionDate
                # ws.cell(row=x+1 , column=9).value = CommonFunctions.ExecutionTime
                ws.cell(row=x + 1, column=9).value = str(time.strftime("%H:%M:%S", time.localtime()))
                ProcessingTime = float(str((time.process_time() - starttime + 2)))
                ws.cell(row=x + 1, column=11).value = ProcessingTime
                ws.cell(row=x + 1, column=13).value = CommonFunctions.SystemUser
                ws.cell(row=x + 1, column=14).value = CommonFunctions.WindowServer

                if (status == 'Passed'):

                    ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                     fill_type='solid')
                else:
                    ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                     fill_type='solid')
                ws.cell(row=x + 1, column=19).value = status
                ws.cell(row=x + 1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                            wrap_text=True, wrapText=True)
                ws.cell(row=x + 1, column=18).value = showcode
                ws.cell(row=x + 1, column=16).value = str(resp)
                wb.save('' + CommonFunctions.OutPutFilePath + '')

        print("-------------------Test Results------------------")
        print("Test Case ID : " + TestCaseID)
        print("URL  : " + URL)
        print("Header  : ")
        print(Parameters)

        print("Test Status  : " + status)
        print("Response  : ")
        print(resp)
        print("--------------------------------------------")

    def UpdateExcelTestCaseproject(self, Sheetname, TestCaseID, URL, Parameters, status, starttime, resp,resp_data):

        showcode = str(resp)

        wb = load_workbook('' + CommonFunctions.OutPutFilePath + '')
        wb.sheetnames
        ws = wb[Sheetname]

        first_column = ws['B']
        del Parameters["AuthToken"]
        for x in range(len(first_column)):
            if (first_column[x].value) == TestCaseID:

                ws.cell(row=x + 1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True, wrapText=True)
                ws.cell(row=x + 1, column=7).value = 'Header \n' + str(Parameters).replace(',', '\n')
                # ws.cell(row=x + 1, column=5).value = URL
                ws.cell(row=x + 1, column=8).value = CommonFunctions.ExecutionDate
                # ws.cell(row=x+1 , column=9).value = CommonFunctions.ExecutionTime
                ws.cell(row=x + 1, column=9).value = str(time.strftime("%H:%M:%S", time.localtime()))
                ProcessingTime = float(str((time.process_time() - starttime + 2)))
                ws.cell(row=x + 1, column=11).value = ProcessingTime
                ws.cell(row=x + 1, column=13).value = CommonFunctions.SystemUser
                ws.cell(row=x + 1, column=14).value = CommonFunctions.WindowServer

                if (status == 'Passed'):

                    ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                     fill_type='solid')
                else:
                    ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                     fill_type='solid')
                ws.cell(row=x + 1, column=19).value = status
                ws.cell(row=x + 1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                            wrap_text=True, wrapText=True)
                ws.cell(row=x + 1, column=18).value = showcode
                ws.cell(row=x + 1, column=16).value = str(resp_data)
                wb.save('' + CommonFunctions.OutPutFilePath + '')

        print("-------------------Test Results------------------")
        print("Test Case ID : " + TestCaseID)
        print("URL  : " + URL)
        print("Header  : ")
        print(Parameters)

        print("Test Status  : " + status)
        print("Response  : ")
        print(resp)
        print("--------------------------------------------")


if __name__ == '__main__':
    obj= CommonFunctions()
    obj.GenerateValidExtension()
        