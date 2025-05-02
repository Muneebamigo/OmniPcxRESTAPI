'''
Created on Jul 19, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time,requests
from Settings import CommonFunctions as CF
from unittest import TestCase
from openpyxl import load_workbook
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from test_Provisioning import test_01_SystemSettings as systemsettingsfunctions
from Key import config
#from test_Provisioning import test_15_ServerUserConfiguration as AddServerUser

SheetName=	'14-Server Permissions'

class Test_1_AddServerPermission(TestCase):
    # Url For Add /ServerPermission/Add
    UrlServerPermissionAdd = '/ServerPermission/Add'

    # Start Test Case No 14-01
    def test_01_AddPermission_With_Default_Admin(self, TestCasesStatus=True):
       
        TestCaseID = '14-01'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Adding Method of Server Permission', 'Add Server Permission with Default Admin User')
        
        # System Settings Function calling
        SystemSettingsSunctions = systemsettingsfunctions.test_1_UpdateSystemSettings()
        SystemSettingsSunctions.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        Name = common.GenrateSimpleStringLimit10()
        Description = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'Name': ''+Name+'',
                      'Description': ''+Description+''
                     }
        
        UrlServerPermissionAdd = '/ServerPermission/Add'
        URL = ''+common.Domain+''+UrlServerPermissionAdd+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                
        else:
            TestCasesStatus=False 
            # Test Case End
        return Name
    
    # Start Test Case No 14-02
    def test_02_AddPermission_With_Default_Admin(self, TestCasesStatus=True):
      
        TestCaseID = '14-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Adding Method of Server Permission', 'Add Server Permission with Default Admin User when server role as secondary configured.')
        
        # System Settings Function calling
        SystemSettingsSunctions = systemsettingsfunctions.test_1_UpdateSystemSettings()
        SystemSettingsSunctions.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        Name = common.GenrateSimpleStringLimit10()
        Description = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'Name': ''+Name+'',
                      'Description': ''+Description+''
                     }
        
        UrlServerPermissionAdd = '/ServerPermission/Add'
        URL = ''+common.Domain+''+UrlServerPermissionAdd+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                
        else:
            TestCasesStatus=False 
            # Test Case End
    
    # Start Test Case No 14-03
    def test_04_AddDuplicatePermission_With_Default_Admin(self, TestCasesStatus=True):
       
        TestCaseID = '14-03'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Adding Method of Server Permission', 'Add Duplicate Server Permission with default admin')
        
        # System Settings Function calling
        SystemSettingsSunctions = systemsettingsfunctions.test_1_UpdateSystemSettings()
        SystemSettingsSunctions.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("SELECT sg_name FROM OPR_Security_Group ORDER BY RANDOM() LIMIT 1;")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_Name = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        Description = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'Name': ''+SG_Name+'',
                      'Description': ''+Description+''
                     }
        
        URL = ''+common.Domain+''+self.UrlServerPermissionAdd+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 409:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        
        else:
            TestCasesStatus=False 
            # Test Case End
            
    # Start Test Case No 14-09
    def test_10_AddPermission_With_Default_Admin(self, TestCasesStatus=True):
       
        TestCaseID = '14-09'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Adding Method of Server Permission', 'Add Server Permission with NULL Name')
        
        # System Settings Function calling
        SystemSettingsSunctions = systemsettingsfunctions.test_1_UpdateSystemSettings()
        SystemSettingsSunctions.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        Name = ''
        Description = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'Name': ''+Name+'',
                      'Description': ''+Description+''
                     }
        
        UrlServerPermissionAdd = '/ServerPermission/Add'
        URL = ''+common.Domain+''+UrlServerPermissionAdd+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                
        else:
            TestCasesStatus=False 
            # Test Case End
        return Name
    
    # Start Test Case No 14-10
    def test_11_AddPermission_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '14-10'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Adding Method of Server Permission', 'Add Server Permission with invalid Character in Name')
        
        # System Settings Function calling
        SystemSettingsSunctions = systemsettingsfunctions.test_1_UpdateSystemSettings()
        SystemSettingsSunctions.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        Name = 'abcB#$@!&*'
        Description = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'Name': ''+Name+'',
                      'Description': ''+Description+''
                     }
        
        UrlServerPermissionAdd = '/ServerPermission/Add'
        URL = ''+common.Domain+''+UrlServerPermissionAdd+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                
        else:
            TestCasesStatus=False 
            # Test Case End
    
    # Start Test Case No 14-11
    def test_12_AddPermission_With_Default_Admin(self, TestCasesStatus=True):
       
        TestCaseID = '14-11'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Adding Method of Server Permission', 'Add Server Permission when server role as branch recorder')
        
        # System Settings Function calling
        SystemSettingsSunctions = systemsettingsfunctions.test_1_UpdateSystemSettings()
        SystemSettingsSunctions.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        Name = common.GenrateSimpleStringLimit10()
        Description = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'Name': ''+Name+'',
                      'Description': ''+Description+''
                     }
        
        UrlServerPermissionAdd = '/ServerPermission/Add'
        URL = ''+common.Domain+''+UrlServerPermissionAdd+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                
        else:
            TestCasesStatus=False 
            # Test Case End
            
class Test_2_GetServerPermission(TestCase):
        #URL for Get ServerUser
        UrlServerPermissionGetAll  = '/ServerUser/Get'
        # Start Test Case No 14-04
        def test_05_GetPermission_With_Default_Admin(self):
           
            TestCaseID = '14-04'
            # Calling Common Functions
            common = CF.CommonFunctions()
            common.Header('Server Permissions', 'Calling Get Method of Server Permission', 'Get All Server Permission with Default Admin User')
            # Test Case Start Time
            starttime = time.process_time()
            # Header Parameters of Rest API
            Parameters = {'AuthToken':config.sessionkey,
                          'AuthUser':config.auth_user
                         }
            
            URL = ''+common.Domain+''+self.UrlServerPermissionGetAll+''
            # Hit API Through Methods
            response = requests.get(URL, headers=Parameters)
            # API Response in JSon Format
            resp = response.json()
            showcode = str(resp['ResponseCode'])
            # Response Code Verification
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                            # Test Case End
        
        # Start Test Case No 14-05
        def test_06_GetPermission_With_Custom_NoAdmin(self):
            # Test Case Start Time
            starttime = time.process_time()
            TestCaseID = '14-05'
            # Calling Common Functions
            common = CF.CommonFunctions()
            common.Header('Server Permissions', 'Calling Get Method of Server Permission', 'Get All Server Permission with Custom User having no server admin perimssions')
            
            ServerAuthUser ='test'
            # Header Parameters of Rest API
            Parameters = {'AuthToken':config.sessionkey,
                          'AuthUser':''+ServerAuthUser+''
                         }
            
            URL = ''+common.Domain+''+self.UrlServerPermissionGetAll+''
            # Hit API Through Methods
            response = requests.get(URL, headers=Parameters)
            # API Response in JSon Format
            resp = response.json()
            showcode = str(resp['ResponseCode'])
            
            # Response Code Verification
            try:
                    if resp['ResponseCode'] == 401:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
                            # Test Case End

class Test_3_DeleteServerPermission(TestCase):
    #URl for Delete Server Permission
    UrlServerPermissionDeleteById  = '/ServerPermission/Delete/'
    # Start Test Case No 14-06
    def test_07_DeletePermission_With_Default_Admin(self):
       
        TestCaseID = '14-06'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Delete Method of Server Permission', 'Delete Server Permission By Id with default admin')
        Name=Test_1_AddServerPermission.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        SQLCommand = ("Select SG_ID from OPR_Security_Group  WHERE SG_Name = '"+Name+"'")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user
                     }
        
        URL = ''+common.Domain+''+self.UrlServerPermissionDeleteById + SG_ID
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['14-Server Permissions']
                first_column = ws['B']
                
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End
    
    # Start Test Case No 14-07
    def test_08_DeletePermission_With_Custom_NoAdmin(self):
       
        TestCaseID = '14-07'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Delete Method of Server Permission', 'Delete Server Permission By Id with Custom user having no admin rights')
        
        Name=Test_1_AddServerPermission.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+Name+"'")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        
        # Test Case Start Time
        starttime = time.process_time()
        ServerAuthUser ='test'
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':''+ServerAuthUser+''
                     }
        
        URL = ''+common.Domain+''+self.UrlServerPermissionDeleteById + SG_ID
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 401:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['14-Server Permissions']
                first_column = ws['B']
                
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
      
    # Start Test Case No 14-08
    def test_09_DeletePermission_With_Default_Admin_InvalidID(self):
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '14-08'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Permissions', 'Calling Delete Method of Server Permission', 'Delete Server Permission By invalid Id with default admin')
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user
                     }
        Invalid_Id ='123456'
        URL = ''+common.Domain+''+self.UrlServerPermissionDeleteById + Invalid_Id
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['14-Server Permissions']
                first_column = ws['B']
                
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        
                        # Test Case End