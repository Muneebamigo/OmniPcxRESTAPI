'''
Created on Aug 1, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time, requests
from Settings import CommonFunctions as CF
from openpyxl import load_workbook
from unittest import TestCase
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from test_Provisioning import test_39_AddCalls
from Key import config
SheetName = '42-Notes'


class test_1_UpdateCallNotes(TestCase):
    UrlForUpdateCallNotes = '/Calls/AddNotes'

    def testcase_01_UpdateCallNotes(self, TestCasesStatus=True):

        TestCaseID = '42-01'
        common = CF.CommonFunctions()
        common.Header('Notes', 'Using Put Method Update Call Notes', 'Update Call Notes.')

        # Calling Add Calls Functions
        calls = test_39_AddCalls.test_1_AddCalls()
        # GlobalCallID, CorrelatorID, PBXCallID, Device = calls.testcase_01_AddCalls(common.PrereqTestCasesStatusUpdate)
        GlobalCallID, CorrelatorID, PBXCallID, Device = calls.testcase_42_AddCalls(common.PrereqTestCasesStatusUpdate)
        GlobalCallID = GlobalCallID
        PBXCallID = PBXCallID
        Device = Device
        # Tenant DB Connectivity Function calling
        cursor = common.DBConnectivity()
        SQLCommand = ("Select CallDetailId From CallDetails Where CorrelatorId  = '" + CorrelatorID + "';")
        cursor.execute(SQLCommand)
        calldetailid = cursor.fetchone()
        CallDetailId = str(calldetailid[0])

        # Header Parameters of Rest API
        DBRecordID = CallDetailId
        NotesData = common.GenrateSimpleStringLimit10()
        SiteCode = ''

        starttime = time.process_time()
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'DBRecordID': '' + DBRecordID + '',
                      'SiteCode': '' + SiteCode + '',
                      'NotesData': '' + NotesData + '',

                      }

        UrlForUpdateCallNotes = '/Calls/AddNotes'
        URL = '' + common.Domain + '' + UrlForUpdateCallNotes + ''
        response = requests.put(URL, headers=Parameters)
        resp = response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False

        return CallDetailId

    def testcase_02_UpdateCallNotes(self, TestCasesStatus=True):

        TestCaseID = '42-02'
        common = CF.CommonFunctions()
        common.Header('Notes', 'Using Put Method Update Call Notes', 'Update Call Notes with invalid DBRecordID.')

        # Header Parameters of Rest API
        DBRecordID = ''
        NotesData = common.GenrateSimpleStringLimit10()
        SiteCode = ''

        starttime = time.process_time()
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'DBRecordID': '' + DBRecordID + '',
                      'SiteCode': '' + SiteCode + '',
                      'NotesData': '' + NotesData + '',

                      }

        URL = '' + common.Domain + '' + self.UrlForUpdateCallNotes + ''
        response = requests.put(URL, headers=Parameters)
        resp = response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False

    def testcase_03_UpdateCallNotes(self, TestCasesStatus=True):

        TestCaseID = '42-03'
        common = CF.CommonFunctions()
        common.Header('Notes', 'Using Put Method Update Call Notes', 'Update Call with Null Notes.')

        # Header Parameters of Rest API
        DBRecordID = '31072018151022666000002'
        NotesData = ''
        SiteCode = ''

        starttime = time.process_time()
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'DBRecordID': '' + DBRecordID + '',
                      'SiteCode': '' + SiteCode + '',
                      'NotesData': '' + NotesData + '',

                      }

        URL = '' + common.Domain + '' + self.UrlForUpdateCallNotes + ''
        response = requests.put(URL, headers=Parameters)
        resp = response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False


class test_2_GetNotes(TestCase):

    def testcase_04_GetNotes(self, TestCasesStatus=True):

        TestCaseID = '42-04'
        common = CF.CommonFunctions()

        # Calling Add Note Functions
        CallDetailId = test_1_UpdateCallNotes.testcase_01_UpdateCallNotes(common.PrereqTestCasesStatusUpdate)

        DBRecordID = CallDetailId

        starttime = time.process_time()
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'SiteCode': '',
                      'DBRecordID': '' + DBRecordID + '',
                      }
        UrlForGetNotes = '/Calls/GetNotes/'

        URL = '' + common.Domain + '' + UrlForGetNotes
        response = requests.get(URL, headers=Parameters)
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        status = 'Failed'
        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            finally:
                wb = load_workbook('' + common.OutPutFilePath + '')
                wb.sheetnames
                ws = wb['42-Notes']
                del Parameters["AuthToken"]
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value) == TestCaseID:
                        ws.cell(row=x + 1, column=7).alignment = XLStyle.Alignment(horizontal='center',
                                                                                   vertical='center', wrap_text=True,
                                                                                   wrapText=True)
                        ws.cell(row=x + 1, column=7).value = str(Parameters).replace(',', '\n')
                        ws.cell(row=x + 1, column=8).value = common.ExecutionDate
                        ws.cell(row=x + 1, column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x + 1, column=11).value = ProcessingTime
                        ws.cell(row=x + 1, column=13).value = common.SystemUser
                        ws.cell(row=x + 1, column=14).value = common.WindowServer
                        if (status == 'Passed'):
                            ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                             fill_type='solid')
                        else:
                            ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                             fill_type='solid')
                        ws.cell(row=x + 1, column=19).value = status
                        ws.cell(row=x + 1, column=18).alignment = XLStyle.Alignment(horizontal='center',
                                                                                    vertical='center', wrap_text=True,
                                                                                    wrapText=True)
                        ws.cell(row=x + 1, column=18).value = showcode
                        ws.cell(row=x + 1, column=16).value = str(resp)
                        wb.save('' + common.OutPutFilePath + '')
        else:
            TestCasesStatus = False

    def testcase_05_GetNotes(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '42-05'
        common = CF.CommonFunctions()

        DBRecordID = ''

        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'SiteCode': '',
                      'DBRecordID': '' + DBRecordID + '',
                      }
        UrlForGetNotes = '/Calls/GetNotes/'

        URL = '' + common.Domain + '' + UrlForGetNotes
        response = requests.get(URL, headers=Parameters)
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        status = 'Failed'
        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            finally:
                wb = load_workbook('' + common.OutPutFilePath + '')
                wb.sheetnames
                ws = wb['42-Notes']
                del Parameters["AuthToken"]
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value) == TestCaseID:
                        ws.cell(row=x + 1, column=7).alignment = XLStyle.Alignment(horizontal='center',
                                                                                   vertical='center', wrap_text=True,
                                                                                   wrapText=True)
                        ws.cell(row=x + 1, column=7).value = str(Parameters).replace(',', '\n')
                        ws.cell(row=x + 1, column=8).value = common.ExecutionDate
                        ws.cell(row=x + 1, column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x + 1, column=11).value = ProcessingTime
                        ws.cell(row=x + 1, column=13).value = common.SystemUser
                        ws.cell(row=x + 1, column=14).value = common.WindowServer
                        if (status == 'Passed'):
                            ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                             fill_type='solid')
                        else:
                            ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                             fill_type='solid')
                        ws.cell(row=x + 1, column=19).value = status
                        ws.cell(row=x + 1, column=18).alignment = XLStyle.Alignment(horizontal='center',
                                                                                    vertical='center', wrap_text=True,
                                                                                    wrapText=True)
                        ws.cell(row=x + 1, column=18).value = showcode
                        ws.cell(row=x + 1, column=16).value = str(resp)
                        wb.save('' + common.OutPutFilePath + '')
        else:
            TestCasesStatus = False

    def testcase_06_GetNotes(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '42-06'
        common = CF.CommonFunctions()

        DBRecordID = '31072018154453016000008'
        SiteCode = '100012233'
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'SiteCode': '' + SiteCode + '',
                      'DBRecordID': '' + DBRecordID + '',
                      }
        UrlForGetNotes = '/Calls/GetNotes/'

        URL = '' + common.Domain + '' + UrlForGetNotes
        response = requests.get(URL, headers=Parameters)
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        status = 'Failed'
        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 500:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            finally:
                wb = load_workbook('' + common.OutPutFilePath + '')
                wb.sheetnames
                ws = wb['42-Notes']
                del Parameters["AuthToken"]
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value) == TestCaseID:
                        ws.cell(row=x + 1, column=7).alignment = XLStyle.Alignment(horizontal='center',
                                                                                   vertical='center', wrap_text=True,
                                                                                   wrapText=True)
                        ws.cell(row=x + 1, column=7).value = str(Parameters).replace(',', '\n')
                        ws.cell(row=x + 1, column=8).value = common.ExecutionDate
                        ws.cell(row=x + 1, column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x + 1, column=11).value = ProcessingTime
                        ws.cell(row=x + 1, column=13).value = common.SystemUser
                        ws.cell(row=x + 1, column=14).value = common.WindowServer
                        if (status == 'Passed'):
                            ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                             fill_type='solid')
                        else:
                            ws.cell(row=x + 1, column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                             fill_type='solid')
                        ws.cell(row=x + 1, column=19).value = status
                        ws.cell(row=x + 1, column=18).alignment = XLStyle.Alignment(horizontal='center',
                                                                                    vertical='center', wrap_text=True,
                                                                                    wrapText=True)
                        ws.cell(row=x + 1, column=18).value = showcode
                        ws.cell(row=x + 1, column=16).value = str(resp)
                        wb.save('' + common.OutPutFilePath + '')
        else:
            TestCasesStatus = False
