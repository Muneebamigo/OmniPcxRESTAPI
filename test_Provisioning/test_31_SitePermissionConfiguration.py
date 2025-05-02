'''
Created on Jul 27, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time, requests
from Settings import CommonFunctions as CF
from openpyxl import load_workbook
from unittest import TestCase
from test_Provisioning import test_01_SystemSettings as SSF
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from Key import config
SheetName=	'31-Site Permissions'

class test_1_AddSitePermissionConfiguration(TestCase):
    
    # Start Test Case No 31-01
    def testcase_01_AddSitePermission (self, TestCasesStatus=True): 
          
        TestCaseID = '31-01'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method Add Site Permission ')
        
        # System Settings Function calling
        SSFunctions=SSF.test_1_UpdateSystemSettings()
        SSFunctions.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Generate Simple Character String Limit 10 Characters
        Name10Char = common.GenrateSimpleStringLimit10()
        Description10Char = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time() 
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description10Char+'',
                          
                    }
        # Url For Add Site Permission 
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 200:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
                
        return Name10Char

    # Test Case End
    
    # Start Test Case No 31-02
    def testcase_1_AddSitePermission(self, TestCasesStatus=True): 
        
        TestCaseID = '31-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method Add Site Permission with non existing Site Code')
        
        # System Settings Function calling
        SSFunctions=SSF.test_1_UpdateSystemSettings()
        SSFunctions.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Generate Simple Character String Limit 10 Characters
        Name10Char = common.GenrateSimpleStringLimit10()
        Description10Char = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'1234567',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description10Char+'',
                          
                    }
        # Url For Add Site Permission 
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 500:
                        print(common.SuccessMessage)
                        status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
    # Test Case End
    
    # Start Test Case No 31-03           
    def testcase_2_AddSitePermission (self, TestCasesStatus=True): 
        
        TestCaseID = '31-03' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method ADD Site Permission with Existing Name ')
        
        # Site Permission Configuration Function calling
        Name10Char = test_1_AddSitePermissionConfiguration.testcase_01_AddSitePermission(common.PrereqTestCasesStatusUpdate)
        
        # Generate Simple Character String Limit 10 Characters
        Description10Char = common.GenrateSimpleStringLimit10()  
        # Test Case Start Time
        starttime = time.process_time()   
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description10Char+'',
                       
                    }
        
        # Url For Add Site Permission   
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        status = 'Failed'
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 409:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
        
    # Test Case End
    
    # Start Test Case No 31-07         
    def testcase_3_AddSitePermission (self, TestCasesStatus=True): 
        
        TestCaseID = '31-07' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method ADD Site Permission with Null Name ')
        
        # Site Permission Configuration Function calling
        Name10Char = ''

        # Generate Simple Character String Limit 10 Characters
        Description10Char = common.GenrateSimpleStringLimit10()  
        # Test Case Start Time
        starttime = time.process_time()   
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description10Char+'',
                       
                    }
        
        # Url For Add Site Permission   
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        status = 'Failed'
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
        
    # Test Case End
    
    # Start Test Case No 31-08          
    def testcase_4_AddSitePermission (self, TestCasesStatus=True): 
        
        TestCaseID = '31-08' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method ADD Site Permission with with description more then 255 characters.')
        
        # Site Permission Configuration Function calling
        Name10Char = common.GenrateSimpleStringLimit10()  
        # Generate Simple Character String Limit 10 Characters
        Description251Char = common.GenrateDesc250()   
        # Test Case Start Time
        starttime = time.process_time()  
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description251Char+'',
                       
                    }
        
        # Url For Add Site Permission   
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        status = 'Failed'
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
        
    # Test Case End
    
    # Start Test Case No 31-09
    def testcase_05_AddSitePermission (self, TestCasesStatus=True): 
        
        TestCaseID = '31-09'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method Add Site Permission when server role as branch recorder')
        
        # System Settings Function calling
        SSFunctions=SSF.test_1_UpdateSystemSettings()
        SSFunctions.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Generate Simple Character String Limit 10 Characters
        Name10Char = common.GenrateSimpleStringLimit10()
        Description10Char = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description10Char+'',
                          
                    }
        # Url For Add Site Permission 
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
                
    # Start Test Case No 31-10
    def testcase_10_AddSitePermission (self, TestCasesStatus=True): 
        
        TestCaseID = '31-10'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Post/Add Site Permissions','SitePermissionConfiguration', 'Using Post Method Add Site Permission when server role as secondary configured.')
        
        # System Settings Function calling
        SSFunctions=SSF.test_1_UpdateSystemSettings()
        SSFunctions.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Generate Simple Character String Limit 10 Characters
        Name10Char = common.GenrateSimpleStringLimit10()
        Description10Char = common.GenrateSimpleStringLimit10()
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
                      'Name': ''+Name10Char+'',
                      'Description': ''+Description10Char+'',
                          
                    }
        # Url For Add Site Permission 
        UrlforAddSitePermission = '/SitePermission/Add'
        URL = ''+common.Domain+''+UrlforAddSitePermission+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification   
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False                      
                
class test_2_GETSitePermission (TestCase):
    
    # Url For Get Site Permission  
    UrlForGetAllSitePermission = '/SitePermission/Get'
    
    # Start Test Case No 31-04
    def testcase_01_GETCustomFields(self, TestCasesStatus=True):
        
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '31-04'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SitePermission' , 'Using Get Method Get SitePermission Data' , 'Get all Data of SitePermission.')
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':'',
    
                    }
        
        # Url For Get Site Permission
        URL = ''+common.Domain+''+self.UrlForGetAllSitePermission+''
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    if resp['list'] == '[]':
                        print(common.SuccessMessage)
                        status = 'Passed - But List is Empty'
                    else:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['31-Site Permissions']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
    
    # Test Case End
                            
class test_3_DeleteSitePermission(TestCase):
    
    UrlDeleteSitePermission = '/SitePermission/Delete/'
    
    # Start Test Case No 31-05
    def testcase_01_DeleteSitePermission(self):
        
        TestCaseID = '31-05'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SitePermission', 'Calling Delete Method of SitePermission', 'Delete SitePermission with valid Id')
        cursor = common.DBConnectivity()
        
        # System Settings Function calling
        Name10Char = test_1_AddSitePermissionConfiguration.testcase_01_AddSitePermission(common.PrereqTestCasesStatusUpdate)
                
        SQLCommand = ("Select SG_ID from OPR_Security_Group where SG_Name = '"+Name10Char+"';")
        cursor.execute(SQLCommand)
        sg_id = cursor.fetchone()
        SG_ID = str(sg_id[0])
        cursor.commit()
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'id': ''+SG_ID+'',
                     }
        
        # Url For Delete Site Permission
        URL = ''+common.Domain +''+self.UrlDeleteSitePermission+''+SG_ID+''
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
            if resp['ResponseCode'] == 200:
                print(common.SuccessMessage)    
                status ='Passed'
                    
            else:
                status='Failed'
                
        # Write Output Result in Excel File
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['31-Site Permissions']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
                            
    # Test Case End
    
    # Start Test Case No 31-06
    def testcase_02_DeleteSitePermission(self):
        
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '31-06'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SitePermission', 'Calling Delete Method of SitePermission', 'Delete SitePermission with valid Id')
        
        SG_ID = '123456'
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'id': ''+SG_ID+'',
                     }
        
        # Url For Delete Site Permission
        URL = ''+common.Domain +''+self.UrlDeleteSitePermission+''+SG_ID+''
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
            if resp['ResponseCode'] == 400:
                print(common.SuccessMessage)    
                status ='Passed'
                    
            else:
                status='Failed'
                
        # Write Output Result in Excel File
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['31-Site Permissions']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
                            
    # Test Case End