'''
Created on Jul 24, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time,requests
from Settings import CommonFunctions as CF
from unittest import TestCase
from openpyxl import load_workbook
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from test_Provisioning import test_01_SystemSettings as SS
from random import randint
from Key import config

SheetName=	'22-SIP Trunk'
# Generate Random DID
class RandomNumbers():
    
    def random_DID(self,n):
        range_start = 10**(n-1)
        range_end = (10**n)-1
        
        return randint(range_start, range_end)

class Add_1_SIPTrunkConfiguration(TestCase):
    
    # Start Test Case No 22-01
    def test_01_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-01'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
        
        # Test Case Start Time    
        starttime = time.process_time()
        # Generate Simple Character String Limit 10 Characters
        SIPTrunkName = common.GenrateSimpleStringLimit10()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':''+SIPTrunkName+'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            
        return SIPTrunkName
        # Test Case End
    
    # Start Test Case No 22-02
    def test_02_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration when server role as branch configured')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
        
        # Generate Simple Character String Limit 10 Characters
        SIPTrunkName = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time    
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':''+SIPTrunkName+'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
    
    # Start Test Case No 22-03
    def test_03_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-03'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration with DID value as null')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Test Case Start Time    
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':''+common.GenrateSimpleStringLimit10()+'',
                      'DID': '',
                      'BranchID': ''
                     }
        
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End
                   
    # Start Test Case No 22-04
    def test_04_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-04'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration with Name value as null')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
      
        # Test Case Start Time    
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
             
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End
            
            
    # Start Test Case No 22-05
    def test_05_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-05'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration with Duplicate Name')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # SIP Trunk Configuration Function calling
        SIPTrunkName = Add_1_SIPTrunkConfiguration.test_01_AddSipTrunk(common.PrereqTestCasesStatusUpdate)
        
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
      
        # Test Case Start Time   
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':''+SIPTrunkName+'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 409:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End
            
    # Start Test Case No 22-06
    def test_06_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-06'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration with Duplicate DID')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # SQL Queries for Data Verification
        'Selecting Random Exisitng security group row from below SQL Query'
        SQLCommand = ("SELECT DDI FROM SIPTrunk ORDER BY RANDOM() LIMIT 1;")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        DID = str(vals[0])
      
        # Test Case Start Time   
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':''+common.GenrateSimpleStringLimit10()+'',
                      'DID': ''+DID+'',
                      'BranchID': ''
                     }
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 409:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End
            
    # Start Test Case No 22-15
    def test_15_AddSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-15'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Add Method of SIP Trunk', 'Add SIP Trunk Configuration when server role as secondary configured')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
        
        # Generate Simple Character String Limit 10 Characters
        SIPTrunkName = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time    
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Name':''+SIPTrunkName+'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        
        # Url For Add SIP Trunk
        URLAddSipTrunk='/SIPTrunk/Add'
        URL = ''+common.Domain+''+URLAddSipTrunk+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            
class Update_2_SIPTrunkConfiguration(TestCase):
    
    # Url For Update SIP Trunk
    UrlUpdateSIPTrunk='/SIPTrunk/Update/'
    
    # Start Test Case No 22-07
    def test_07_UpdateSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-07'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Update Method of SIP Trunk', 'Update SIP Trunk Configuration')
        
        # SIP Trunk Configuration Function calling
        addSIPTrunk = Add_1_SIPTrunkConfiguration()
        SIPtrunkName = addSIPTrunk.test_01_AddSipTrunk(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # SQL Queries for Data Verification
        SQLCommand = ("  SELECT Id from SIPTrunk where Name = '"+SIPtrunkName+"'")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        Id = str(vals[0])
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
      
        # Test Case Start Time   
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID':''+Id+'',
                      'Name':''+common.GenrateSimpleStringLimit10()+'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        # Url For Update SIP Trunk
        URL = ''+common.Domain+''+self.UrlUpdateSIPTrunk+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End

    # Start Test Case No 22-08
    def test_08_UpdateSipTrunk(self, TestCaseStatus=True):
        
        TestCaseID = '22-08'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Update Method of SIP Trunk', 'Update SIP Trunk Configuration With Empty Name Parameter Value')
   
        # SIP Trunk Configuration Function calling
        addSIPTrunk = Add_1_SIPTrunkConfiguration()
        SIPtrunkName = addSIPTrunk.test_01_AddSipTrunk(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # SQL Queries for Data Verification
        SQLCommand = ("  SELECT Id from SIPTrunk where Name = '"+SIPtrunkName+"'")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        Id = str(vals[0])
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
      
        # Test Case Start Time    
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID':''+Id+'',
                      'Name':'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        # Url For Update SIP Trunk
        URL = ''+common.Domain+''+self.UrlUpdateSIPTrunk+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End

    # Start Test Case No 22-09
    def test_09_UpdateSipTrunk(self, TestCaseStatus=True):
        
        # Test Case Start Time   
        starttime = time.process_time()
        TestCaseID = '22-09'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Update Method of SIP Trunk', 'Update SIP Trunk Configuration with non existing ID')
        
        Id = ''
        # Calling Random DID
        DID = RandomNumbers().random_DID(11)
      
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID':''+Id+'',
                      'Name':''+common.GenrateSimpleStringLimit10()+'',
                      'DID': ''+str(DID)+'',
                      'BranchID': ''
                     }
        # Url For Update SIP Trunk
        URL = ''+common.Domain+''+self.UrlUpdateSIPTrunk+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    wb=load_workbook(''+common.OutPutFilePath+'')
                    wb.sheetnames
                    ws = wb['22-SIP Trunk']
                    first_column = ws['B']
                    
                    del Parameters["AuthToken"]
                    for x in range(len(first_column)):
                            if (first_column[x].value)  == TestCaseID:
                                ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                                ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                                ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                                ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                                ProcessingTime = float(str((time.process_time() - starttime + 2)))
                                ws.cell(row=x+1 , column=11).value = ProcessingTime
                                ws.cell(row=x+1 , column=13).value = common.SystemUser
                                ws.cell(row=x+1 , column=14).value = common.WindowServer
                                if(status =='Passed'):
                                    ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                                else:
                                    ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                                ws.cell(row=x+1 , column=19).value = status
                                ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                                ws.cell(row=x+1 , column=18).value = showcode
                                ws.cell(row=x+1 , column=16).value = str(resp)
                                wb.save(''+common.OutPutFilePath+'')
        else: 
            TestCaseStatus = False
            # Test Case End

class Get_3_SIPTrunkConfiguration(TestCase): 
    
    # Url For Get SIP Trunk
    UrlGetSipTrunk = '/SIPTrunk/Get'
    UrlGetSipTrunkById = '/SIPTrunk/Get/'
    
    # Start Test Case No 22-10
    def test_10_GetSIPTrunkConfiguration(self, TestCaseStatus=True):
        
        TestCaseID = '22-10'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Get Method of SIP Trunk', 'Get List of SIP Trunk Configuration ')
        
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user
                     }
        # Url For Get SIP Trunk
        URL = ''+common.Domain+''+self.UrlGetSipTrunk+''
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End

    
    # Start Test Case No 22-11
    def test_11_GetSIPTrunkConfiguration(self, TestCaseStatus=True):
        
        TestCaseID = '22-11'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Get Method of SIP Trunk', 'Get SIP Trunk Configuration By valid Id')
        
        # SIP Trunk Configuration Function calling
        addSIPTrunk = Add_1_SIPTrunkConfiguration()
        SIPtrunkName = addSIPTrunk.test_01_AddSipTrunk(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # SQL Queries for Data Verification
        SQLCommand = ("SELECT Id from SIPTrunk where Name = '"+SIPtrunkName+"'")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        Id = str(vals[0])
       
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID': ''+Id+'',
                     }
        
        # Url For Get SIP Trunk
        URL = ''+common.Domain+''+self.UrlGetSipTrunkById+Id+''
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File           
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End
            

    # Start Test Case No 22-12   
    def test_12_GetSIPTrunkConfiguration(self, TestCaseStatus=True):
        
        TestCaseID = '22-12'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Get Method of SIP Trunk', 'Get SIP Trunk Configuration By Invalid Id')
       
        # System Settings Function calling
        systemFunc = SS.test_1_UpdateSystemSettings()
        systemFunc.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        Id = '123456'
       
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID': ''+Id+'',
                     }
        
        # Url For Get SIP Trunk
        URL = ''+common.Domain+''+self.UrlGetSipTrunkById+Id+''
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End


class Delete_4_SIPTrunkConfiguration(TestCase):
    
    # Url For Delete SIP Trunk
    UrlDeleteSIPTrunk='/SIPTrunk/Delete/'
    
    # Start Test Case No 22-13
    def test_13_DeleteSIPTrunkConfiguration(self, TestCaseStatus=True):
        
        TestCaseID = '22-13'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Delete Method of SIP Trunk', 'Deleting SIP Trunk by valid ID')
   
        # SIP Trunk Configuration Function calling
        addSIPTrunk = Add_1_SIPTrunkConfiguration()
        SIPtrunkName = addSIPTrunk.test_01_AddSipTrunk(common.PrereqTestCasesStatusUpdate)
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # SQL Queries for Data Verification
        SQLCommand = ("  SELECT Id from SIPTrunk where Name = '"+SIPtrunkName+"'")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        Id = str(vals[0])
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user
                     }
        # Url For Delete SIP Trunk
        URL = ''+common.Domain+''+self.UrlDeleteSIPTrunk+Id+''
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End
                
    # Start Test Case No 22-14
    def test_14_DeleteSIPTrunkConfiguration(self, TestCaseStatus=True):
        
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '22-14'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('SIP Trunk', 'Calling Delete Method of SIP Trunk', 'Deleting SIP Trunk by invalid ID')
   
        Id = '123456'
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'id': ''+Id+'',
                     }
        # Url For Delete SIP Trunk
        URL = ''+common.Domain+''+self.UrlDeleteSIPTrunk+Id+''
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCaseStatus == True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else: 
            TestCaseStatus = False
            # Test Case End