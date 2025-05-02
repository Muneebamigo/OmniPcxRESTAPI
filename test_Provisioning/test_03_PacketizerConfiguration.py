'''
Created on Jul 4, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed


----------------pre requisite--------------------------------------
System setting must be configured.
Node mean Pbx also added
Node or PBx are added from pbx configuration


------------------OutPut----------------------------
This module will be add packetizer if node is added other wise its not able to add it.
by update request packetizer settings is updated and it will be shown on Packetizer settings page of the server Administration.
Addded  packetizer also be delete , remove Packetizer settings page of the server Administration.
All the packetizer also get which are avaliable on  Packetizer settings page of the server Administration.
'''

import time, requests, random
from Settings import CommonFunctions as CF
from openpyxl import load_workbook
from unittest import TestCase
from test_Provisioning import test_01_SystemSettings as systemsettingsFunctions
from test_Provisioning import test_19_Branch as branchFunctions
from test_Provisioning import test_02_PBXConfiguration as pbxconfigurationFunctions
from test_Provisioning import test_04_DeviceConfigurations as DCF
from test_Provisioning import test_36_Board as BoardFun
from test_Provisioning import test_38_TrunkGroup as TGF
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from Key import config
SheetName=    '3-Packetizer Configuration'

class test_1_AddPacketizerConfiguration(TestCase):
    
    # Start Test Case No 03-01
    def testcase_01_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-01'
        pbxfunctionscall = pbxconfigurationFunctions.test_1_AddPBXConfiguration()

        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Packetizer When Recorder Type is Extension With Valid Node ID.')



        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()

        # System Settings Function calling
        SystemSettingsFunction = systemsettingsFunctions.test_1_UpdateSystemSettings()
        SystemSettingsFunction.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)

        #PBX Configuration Functions Calling
        OXEName, ValidIP=pbxfunctionscall.testcase_01_AddPBXConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID From PBXDetail Where OXEName = '"+OXEName+"' and OXEIP = '"+ValidIP+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        PBXID=str(pbxid[0])
        cursor.commit()
        
        # Generate Valid IP
        PcktIP=common.GenrateValidIPString()
        
        # Generate Simple Character String Limit 10 Character
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        
        # Url For Add Packetizer 
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # SQL Queries for Data Verification
        SQLCommand1 = ("Select IP from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand1)
        packetizerip=cursor.fetchone()
        cursor.commit()
        
        # Response Code Verification
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print('a')
                    if packetizerip[0] == PcktIP:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File       
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
           
        return PcktIP, PBXID
    # Test Case End
    
    # Start Test Case No 03-02
    def testcase_02_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Extension With InValid Node ID.')
        
        PcktIP=common.GenrateValidIPString()
        PBXID = '12345'
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    # Start Test Case No 03-03
    def testcase_03_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-03'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Extension With Duplicate IP Address.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        pbxfunctionscall = pbxconfigurationFunctions.test_1_AddPBXConfiguration()
        OXEName, ValidIP=pbxfunctionscall.testcase_01_AddPBXConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID From PBXDetail Where OXEName = '"+OXEName+"' and OXEIP = '"+ValidIP+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        PBXID=str(pbxid[0])
        cursor.commit()
        
        PcktIP=common.GenrateValidIPString()
        crystal1 = str(random.randint(1 , 99))
        crystal2 = str(random.randint(1 , 99))
        coupler1 = str(random.randint(1 , 99))
        coupler2 = str(random.randint(1 , 99))
        port1= str(random.randint(1000 , 99999))
        port2= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        Parameters1 = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal1+'',
                      'Coupler': ''+coupler1+'',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port1+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        Parameters2 = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal2+'',
                      'Coupler': ''+coupler2+'',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port2+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        requests.post(URL, headers=Parameters1)
        response2 = requests.post(URL, headers=Parameters2)
        resp2=response2.json()
        showcode = str(resp2['ResponseCode'])
        
        
        if TestCasesStatus==True:
            try:
                if resp2['ResponseCode'] == 409:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['3-Packetizer Configuration']
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1, column=7).value = 'Pecketizer1 Data:: \n RecorderType= 0 \n PBXID= '+PBXID+'  \n Crystal= '+crystal1+'  \n Coupler= '+coupler1+'  \n IPAddress= '+PcktIP+'  \n Port= '+port1+' \n BranchID=  \n\n Pecketizer2 Data:: \n RecorderType= 0/Extension \n PBXID= '+PBXID+'  \n Crystal= '+crystal2+'  \n Coupler= '+coupler2+'  \n IPAddress= '+PcktIP+'  \n Port= '+port2+' \n BranchID=  \n '
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp2)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Start Test Case No 03-04
    def testcase_04_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-04'
        pbxfunctionscall = pbxconfigurationFunctions.test_1_AddPBXConfiguration()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Extension With Duplicate Crystal & Coupler.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        OXEName, ValidIP=pbxfunctionscall.testcase_01_AddPBXConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID From PBXDetail Where OXEName = '"+OXEName+"' and OXEIP = '"+ValidIP+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        PBXID=str(pbxid[0])
        cursor.commit()
        
        PcktIP1=common.GenrateValidIPString()
        PcktIP2=common.GenrateValidIPString()
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        port1= str(random.randint(1000 , 99999))
        port2= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        Parameters1 = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIP1+'',
                      'Port': ''+port1+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        Parameters2 = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIP2+'',
                      'Port': ''+port2+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        requests.post(URL, headers=Parameters1)
        response2 = requests.post(URL, headers=Parameters2)
        resp2=response2.json()
        showcode = str(resp2['ResponseCode'])
        
        
        if TestCasesStatus==True:
            try:
                if resp2['ResponseCode'] == 409:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['3-Packetizer Configuration']
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1, column=7).value = 'Pecketizer1 Data:: \n RecorderType= 0 \n PBXID= '+PBXID+'  \n Crystal= '+crystal+'  \n Coupler= '+coupler+'  \n IPAddress= '+PcktIP1+'  \n Port= '+port1+' \n BranchID=  \n\n Pecketizer2 Data:: \n RecorderType= 0/Extension \n PBXID= '+PBXID+'  \n Crystal= '+crystal+'  \n Coupler= '+coupler+'  \n IPAddress= '+PcktIP2+'  \n Port= '+port2+' \n BranchID=  \n '                                          
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp2)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Start Test Case No 03-05
    def testcase_05_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-05'
        SystemSettingsFunction = systemsettingsFunctions.test_1_UpdateSystemSettings()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Trunk.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        # SystemSettingsFunction.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        PcktIP=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '1',
                      'PBXID': '',
                      'Crystal': '',
                      'Coupler': '',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # SQL Queries for Data Verification
        SQLCommand1 = ("Select IP from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand1)
        packetizerip=cursor.fetchone()
        cursor.commit()
        
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print('a')
                    if packetizerip[0] == PcktIP:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
        return PcktIP
    
    # Start Test Case No 03-06
    def testcase_06_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-06'
        BranchFunction = branchFunctions.test_1_AddBranch()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Trunk with Valid Branch ID.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        BranchName = BranchFunction.testcase_01_AddBranch(common.PrereqTestCasesStatusUpdate)
        
        SQLCommand = ("Select BranchID From Branch Where BranchName = '"+BranchName+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        BranchID=str(pbxid[0])
        cursor.commit()
        
        PcktIP=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '1',
                      'PBXID': '',
                      'Crystal': '',
                      'Coupler': '',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': ''+BranchID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # SQL Queries for Data Verification
        SQLCommand1 = ("Select IP from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand1)
        packetizerip=cursor.fetchone()
        cursor.commit()
        
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print('a')
                    if packetizerip[0] == PcktIP:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
        return PcktIP,BranchID
    
    # Start Test Case No 03-07
    def testcase_07_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-07'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Trunk with InValid Branch ID.')
        
        BranchID = '12345'
        PcktIP=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '1',
                      'PBXID': '',
                      'Crystal': '',
                      'Coupler': '',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': ''+BranchID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    # Start Test Case No 03-17
    def testcase_17_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-17'
        BranchFunction = branchFunctions.test_1_AddBranch()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Trunk with Duplicate IP.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        BranchName = BranchFunction.testcase_01_AddBranch(common.PrereqTestCasesStatusUpdate)
        
        SQLCommand = ("Select BranchID From Branch Where BranchName = '"+BranchName+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        BranchID=str(pbxid[0])
        cursor.commit()
        
        PcktIP1=common.GenrateValidIPString()    
        port1= str(random.randint(1000 , 99999))
        port2= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters1 = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '1',
                      'PBXID': '',
                      'Crystal': '',
                      'Coupler': '',
                      'IPAddress': ''+PcktIP1+'',
                      'Port': ''+port1+'',
                      'BranchID': ''+BranchID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Header Parameters of Rest API
        Parameters2 = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '1',
                      'PBXID': '',
                      'Crystal': '',
                      'Coupler': '',
                      'IPAddress': ''+PcktIP1+'',
                      'Port': ''+port2+'',
                      'BranchID': ''+BranchID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        requests.post(URL, headers=Parameters1)
        response = requests.post(URL, headers=Parameters2)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 409:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['3-Packetizer Configuration']
                del Parameters2["AuthToken"]
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters2).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
           
    # Start Test Case No 03-19
    def testcase_19_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-19'
        pbxfunctionscall = pbxconfigurationFunctions.test_1_AddPBXConfiguration()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Packetizer When recorder is Main and Server Role configured as secondary.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        #PBX Configuration Functions Calling
        OXEName, ValidIP=pbxfunctionscall.testcase_01_AddPBXConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID From PBXDetail Where OXEName = '"+OXEName+"' and OXEIP = '"+ValidIP+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        PBXID=str(pbxid[0])
        cursor.commit()
        
        #System Settings Configuration Functions Calling
        SystemSettingsFunction = systemsettingsFunctions.test_1_UpdateSystemSettings()
        SystemSettingsFunction.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Generate Valid IP
        PcktIP=common.GenrateValidIPString()
        
        # Generate Simple Character String Limit 10 Characte
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        
        # Url For Add Packetizer 
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 403:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File       
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
           
    # Test Case End
    
    
    # Start Test Case No 03-20
    def testcase_20_AddPacketizerConfiguration(self, TestCasesStatus=True):

        TestCaseID = '03-20'
        pbxfunctionscall = pbxconfigurationFunctions.test_1_AddPBXConfiguration()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer', 'Using Post Method Add Packetizer',
                      'Configure the Packetizer When Recorder is branch.')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()

        # System Settings Function calling
        SystemSettingsFunction = systemsettingsFunctions.test_1_UpdateSystemSettings()
        SystemSettingsFunction.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)

        # PBX Configuration Functions Calling
        OXEName, ValidIP = pbxfunctionscall.testcase_01_AddPBXConfiguration(common.PrereqTestCasesStatusUpdate)


        # SQL Queries for Data Verification
        SQLCommand = ("Select ID From PBXDetail Where OXEName = '" + OXEName + "' and OXEIP = '" + ValidIP + "';")
        cursor.execute(SQLCommand)
        pbxid = cursor.fetchone()
        PBXID = str(pbxid[0])
        cursor.commit()

        # System Settings Configuration Functions Calling
        SystemSettingsFunction = systemsettingsFunctions.test_1_UpdateSystemSettings()
        SystemSettingsFunction.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)

        # Generate Valid IP
        PcktIP = common.GenrateValidIPString()

        # Generate Simple Character String Limit 10 Characte
        crystal = str(random.randint(1, 99))
        coupler = str(random.randint(1, 99))
        port = str(random.randint(1000, 99999))

        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkey,
                      'AuthUser': config.auth_user,
                      'RecorderType': '0',
                      'PBXID': '' + PBXID + '',
                      'Crystal': '' + crystal + '',
                      'Coupler': '' + coupler + '',
                      'IPAddress': '' + PcktIP + '',
                      'Port': '' + port + '',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',

                      }

        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = '' + common.Domain + '' + UrlForAddPacketizer + ''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])

        # Response Code Verification

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 403:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False

            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False

    # Test Case End
    
    
    # Start Test Case No 03-22
    def testcase_22_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-22'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Pecketizer When Recorder Type is Trunk and also add Trunk group and Board.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        BoardFunctions = BoardFun.test_1_AddBoard()
        BoardName= BoardFunctions.testcase_01_AddBoard(common.PrereqTestCasesStatusUpdate)
        # SQL Queries for Data Verification
        SQLCommand = ("Select BoardID From Board Where Name = '"+BoardName+"';")
        cursor.execute(SQLCommand)
        val=cursor.fetchone()
        BoardID=str(val[0])
        cursor.commit()
        
        TrunkGroupFunctions = TGF.test_1_AddTrunkGroup()
        TrunkGroupName, PBXID= TrunkGroupFunctions.testcase_01_AddTrunkGroup(common.PrereqTestCasesStatusUpdate)
        # SQL Queries for Data Verification
        SQLCommand = ("Select TrunkGroupID From TrunkGroup Where Name = '"+TrunkGroupName+"' and PBXID = '"+PBXID+"';")
        cursor.execute(SQLCommand)
        val=cursor.fetchone()
        TrunkGroupID=str(val[0])
        cursor.commit()
        
        PcktIP=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '1',
                      'PBXID': '',
                      'Crystal': '',
                      'Coupler': '',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Channel': '0',
                      'TrunkGroupID': ''+TrunkGroupID+'',
                      'BoardID': ''+BoardID+'',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        # Url For Add Packetizer
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # SQL Queries for Data Verification
        SQLCommand1 = ("Select IP from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand1)
        packetizerip=cursor.fetchone()
        cursor.commit()
        
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print('a')
                    if packetizerip[0] == PcktIP:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
        return PcktIP,BoardID,TrunkGroupID
    
    # Start Test Case No 03-01
    def testcase_23_AddPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-23'
        pbxfunctionscall = pbxconfigurationFunctions.test_1_AddPBXConfiguration()
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Post Method Add Packetizer' , 'Configure the Packetizer When Recorder Type is Extension With Site Session key.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        #PBX Configuration Functions Calling
        OXEName, ValidIP=pbxfunctionscall.testcase_01_AddPBXConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID From PBXDetail Where OXEName = '"+OXEName+"' and OXEIP = '"+ValidIP+"';")
        cursor.execute(SQLCommand)
        pbxid=cursor.fetchone()
        PBXID=str(pbxid[0])
        cursor.commit()
        
        # Generate Valid IP
        PcktIP=common.GenrateValidIPString()
        
        # Generate Simple Character String Limit 10 Character
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIP+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        
        # Url For Add Packetizer 
        UrlForAddPacketizer = '/Packetizer/Add/'
        URL = ''+common.Domain+''+UrlForAddPacketizer+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        
        # Response Code Verification
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 401:
                   
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File       
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

class test_2_UpdatePacketizerConfiguration(TestCase):
    
    # Url For Update Packetizer
    PacketizerUrlForUpdatePBX = '/Packetizer/Update/'
    
    # Start Test Case No 03-08
    def testcase_08_UpdatePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-08'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Put Method Update Packetizer' , 'Update the Pecketizer With Valid Data.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        PcktIP, PBXID = test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()
        PacktID=str(Packtid[0])
        cursor.commit()
        
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        PcktIPp=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIPp+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Id': ''+PacktID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        
        URL = ''+common.Domain+''+self.PacketizerUrlForUpdatePBX+''
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
    
    
    # Start Test Case No 03-09
    def testcase_09_UpdatePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-09'
        
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Put Method Update Packetizer' , 'Update the Pecketizer With InValid Data.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        PcktIP, PBXID= test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()
        PacktID=str(Packtid[0])
        cursor.commit()
        PBXID=PBXID
        
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        PcktIPp='123456'   
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIPp+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Id': ''+PacktID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        
        URL = ''+common.Domain+''+self.PacketizerUrlForUpdatePBX+''
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    # Start Test Case No 03-16
    def testcase_16_UpdatePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-16'
        
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Put Method Update Packetizer' , 'Update the Pecketizer With Valid Data with non existing ID.')
        
        PcktIP, PBXID = test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        PcktIP=PcktIP
        PacktID='123456'
        
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        PcktIPp=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIPp+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Id': ''+PacktID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        PacketizerUrlForUpdatePBX = '/Packetizer/Update/'
        URL = ''+common.Domain+''+PacketizerUrlForUpdatePBX+''
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
            
    def testcase_24_UpdatePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-24'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Put Method Update Packetizer' , 'Update the Pecketizer With Site session key.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        PcktIP, PBXID = test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()
        PacktID=str(Packtid[0])
        cursor.commit()
        
        crystal = str(random.randint(1 , 99))
        coupler = str(random.randint(1 , 99))
        PcktIPp=common.GenrateValidIPString()    
        port= str(random.randint(1000 , 99999))
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'RecorderType': '0',
                      'PBXID': ''+PBXID+'',
                      'Crystal': ''+crystal+'',
                      'Coupler': ''+coupler+'',
                      'IPAddress': ''+PcktIPp+'',
                      'Port': ''+port+'',
                      'BranchID': '',
                      'Id': ''+PacktID+'',
                      'Channel': '',
                      'TrunkGroupID': '',
                      'BoardID': '',
                      'DefaultTrunkChannelEnabled': 'False',
    
                    }
        
        URL = ''+common.Domain+''+self.PacketizerUrlForUpdatePBX+''
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 401:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
                
class test_3_GetPacketizerConfiguration(TestCase):
    
    # Url For Get Packetizer
    UrlForGetAllPacketizerData = '/Packetizer/Get/'
    UrlForGetSinglePacketizerData = '/Packetizer/Get/'
    UrlForGetPacketizerDataByPBXIPORHostName = "/Packetizer/Get/?PBXIPORHostName='Trunk'"
    
    # Start Test Case No 03-10
    def testcase_10_GetPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-10'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Get Method Get Packetizer Data' , 'Get all Data of Pecketizer.')
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForGetAllPacketizerData+''
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    if resp['list'] == '[]':
                        print(common.SuccessMessage)
                        status = 'Passed - But List is Empty'
                    else:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
    
    # Start Test Case No 03-11
    def testcase_11_GetPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-11'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Get Method Get Packetizer Data' , 'Get Specific Data of Pecketizer through Valid ID.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        PcktIP, PBXID = test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()
        PacktID=str(Packtid[0])
        cursor.commit()
        PBXID=PBXID
                
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': ''+PacktID+'',
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForGetSinglePacketizerData+''+PacktID+''
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    if resp['list'][0]['IPAddress'] == PcktIP:
                        print(common.SuccessMessage)
                        status = 'Passed'
                    else:
                        status = 'Failed - But List is empty'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
    
    
    # Start Test Case No 03-12
    def testcase_12_GetPacketizerConfiguration(self, TestCasesStatus=True):
        
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '03-12'
        
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Get Method Get Packetizer Data' , 'Get Specific Data of Pecketizer through InValid ID.')
       
        PacktID = '12345'
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': ''+PacktID+'',
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForGetSinglePacketizerData+''+PacktID+''
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            
            # Write Output Result in Excel File        
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    
    # Start Test Case No 03-13
    def testcase_13_GetPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-13'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Get Method Get Packetizer Data' , 'Get All Data of Pecketizer through Through HostName Trunk.')
        
        PcktIP = test_1_AddPacketizerConfiguration.testcase_05_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        PcktIP = PcktIP  
        
        # Test Case Start Time
        starttime = time.process_time()     
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForGetPacketizerDataByPBXIPORHostName+''
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
        # Start Test Case No 03-18
    def testcase_18_GetPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-18'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Get Method Get Packetizer Data' , 'Get All Data of Pecketizer through Through HostName Trunk is invalid or empty.')
        
        PcktIP = test_1_AddPacketizerConfiguration.testcase_05_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        PcktIP = PcktIP  
        
        # Test Case Start Time
        starttime = time.process_time()     
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
    
                    }
        
        invalidUrlForGetPacketizerDataByPBXIPORHostName = "/Packetizer/Get/?PBXIPORHostName="
        URL = ''+common.Domain+''+invalidUrlForGetPacketizerDataByPBXIPORHostName+''
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    # Start Test Case No 03-25
    def testcase_25_GetPacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-25'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Get Method Get Packetizer Data' , 'Get all Data of Packetizer with site session key.')
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForGetAllPacketizerData+''
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 401:
                    if resp['list'] == '[]':
                        print(common.SuccessMessage)
                        status = 'Passed - But List is Empty'
                    else:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
                    
            
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False 
            
class test_4_DeletePacketizerConfiguration(TestCase):
    
    UrlForDeletePacketizer = '/Packetizer/Delete/'
    
    # Start Test Case No 03-14
    def testcase_14_DeletePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-14'
        common = CF.CommonFunctions()
        
        # Calling Common Functions
        common.Header('Packetizer' , 'Using Delete Method Delete Packetizer' , 'Delete Pecketizer with valid ID.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        PcktIP, PBXID= test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()
        PacktID=str(Packtid[0])
        cursor.commit()
        PBXID=PBXID
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': ''+PacktID+'',
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForDeletePacketizer+''+PacktID+''
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            
            # Write Output Result in Excel File# Write Output Result in Excel File        
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
    
    
    # Start Test Case No 03-15
    def testcase_15_DeletePacketizerConfiguration(self, TestCasesStatus=True):
        
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '03-15'
        
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Packetizer' , 'Using Delete Method Delete Packetizer' , 'Delete Pecketizer with Invalid ID.')
        
        PacktID = '12345'
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': ''+PacktID+'',
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForDeletePacketizer+''+PacktID+''
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            
            # Write Output Result in Excel File        
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            # Test Case End
            
            
    # Start Test Case No 03-21
    def testcase_21_DeletePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-21'
        common = CF.CommonFunctions()
        
        # Calling Common Functions
        common.Header('Packetizer' , 'Using Delete Method Delete Packetizer' , 'Delete Pecketizer with valid ID when packetizer associated with devices.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        DeviceConfigurationFunctions=DCF.test_1_AddDeviceConfiguration()
        ExtVal, PcktIP, PBXID =  DeviceConfigurationFunctions.testcase_40_Add_TDM_Ext_With_ROD_RFN(common.PrereqTestCasesStatusUpdate)
        ExtVal=ExtVal
        PBXID=PBXID
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()

        PacktID=str(Packtid[0])
        cursor.commit()
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': ''+PacktID+'',
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForDeletePacketizer+''+PacktID+''
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 409:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            
            # Write Output Result in Excel File# Write Output Result in Excel File        
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    def testcase_26_DeletePacketizerConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '03-26'
        common = CF.CommonFunctions()
        
        # Calling Common Functions
        common.Header('Packetizer' , 'Using Delete Method Delete Packetizer' , 'Delete Pecketizer with valid ID.')
        # Config DB Connectivity Function calling
        cursor  = common.StringDBConnectivity()
        
        PcktIP, PBXID= test_1_AddPacketizerConfiguration.testcase_01_AddPacketizerConfiguration(common.PrereqTestCasesStatusUpdate)
        
        # SQL Queries for Data Verification
        SQLCommand = ("Select ID from BlueBox Where IP = '"+PcktIP+"';")
        cursor.execute(SQLCommand)
        Packtid=cursor.fetchone()
        PacktID=str(Packtid[0])
        cursor.commit()
        PBXID=PBXID
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'Id': ''+PacktID+'',
    
                    }
        
        URL = ''+common.Domain+''+self.UrlForDeletePacketizer+''+PacktID+''
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 401:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False
            
            # Write Output Result in Excel File# Write Output Result in Excel File        
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            