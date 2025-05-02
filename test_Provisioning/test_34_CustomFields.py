'''
Created on Jul 24, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time, requests, random
from Settings import CommonFunctions as CF
from openpyxl import load_workbook
from unittest import TestCase
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from test_Provisioning import test_01_SystemSettings as systemfuncitons
from Key import config
SheetName=	'34-Custom Fields'

class test_1_UpdateCustomFields(TestCase):  
    
    # Start Test Case No 34-02
    def testcase_01_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Update all Data of CustomFields.')
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Random Integers
        ID= str(random.randint(1, 10))
        # Generate Simple Character String Limit 10 Characters
        SIPTag10Char =common.GenrateSimpleStringLimit10()
        Title10Char = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'False',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'False',
    
                        }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 200:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False    
                
    # Test Case End
    
    # Start Test Case No 34-03
    def testcase_02_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-03' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Enable SIPEnabled.')
        
        # Generate Random Integers
        ID= '1'
        # Generate Simple Character String Limit 10 Characters
        SIPTag10Char ='SIPTAG1'
        Title10Char = 'CusF1'
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'True',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'True',
    
                    }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 200:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
                
        return Title10Char, SIPTag10Char
                
    # Test Case End   
    
    # Start Test Case No 34-04
    def testcase_03_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-04' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Enable SIPEnabled and SIPTag is null')
    
        # Generate Random Integers
        ID= str(random.randint(1, 10))
        # Generate Simple Character String Limit 10 Characters
        Title10Char = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'True',
                      'SIPTag': '',
                      'SIPEnabled': 'True',
    
                        }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False  
                
    # Test Case End  
    
    # Start Test Case No 34-05
    def testcase_04_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-05' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'ID is null.')
    
        # Generate Simple Character String Limit 10 Characters
        SIPTag10Char =common.GenrateSimpleStringLimit10()
        Title10Char = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': "",   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'True',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'True',
    
                        }
        
        # Url For Update Custom Fields   
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False    
                
    # Test Case End
    
    # Start Test Case No 34-06
    def testcase_05_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-06'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Update all Data of CustomFields when server role as branch recorder.')
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Random Integers
        ID= str(random.randint(1, 10))
        # Generate Simple Character String Limit 10 Characters
        SIPTag10Char =common.GenrateSimpleStringLimit10()
        Title10Char = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'False',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'False',
    
                        }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False    
                
    # Test Case End
    
    # Start Test Case No 34-07
    def testcase_07_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-07'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Update all Data of CustomFields with more then 10 Alphabets.')
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Random Integers
        ID= str(random.randint(1, 10))
        # Generate Simple Character String Limit 10 Characters
        SIPTag10Char =common.GenrateSimpleStringLimit10()
        Title10Char = 'abcdefghijklmnopqrstuv'
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'False',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'False',
    
                        }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False    
                
    # Test Case End
    
    # Start Test Case No 34-08
    def testcase_08_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-08'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Update all Data of CustomFields when server role as secondary configured.')
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Random Integers
        ID= str(random.randint(1, 10))
        # Generate Simple Character String Limit 10 Characters
        SIPTag10Char =common.GenrateSimpleStringLimit10()
        Title10Char = common.GenrateSimpleStringLimit10()
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'False',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'False',
    
                        }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False    
                
    # Test Case End
    
    
    # Start Test Case No 34-09
    def testcase_09_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-09' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Enable SIPEnabled and duplicate custom field title.')
        
        # Generate Random Integers
        ID= '2'
        
        Title10Char, SIPTag10Char = test_1_UpdateCustomFields.testcase_02_PUTCustomFields(self,common.PrereqTestCasesStatusUpdate)
        SIPTag10Char=SIPTag10Char
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'True',
                      'SIPTag': 'testtagtest',
                      'SIPEnabled': 'True',
    
                    }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 409:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
                
    # Test Case End 
    
    
    # Start Test Case No 34-10
    def testcase_10_PUTCustomFields(self, TestCasesStatus=True): 
        
        TestCaseID = '34-10' 
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Put Method to Update CustomFields Data' , 'Enable SIPEnabled with duplicate SIP Tag value.')
        
        # Generate Random Integers
        ID= '2'
        
        Title10Char, SIPTag10Char = test_1_UpdateCustomFields.testcase_02_PUTCustomFields(self,common.PrereqTestCasesStatusUpdate)
        Title10Char=Title10Char
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Test Case Start Time
        starttime = time.process_time()   
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode': '',
                      'ID': ''+ID+'',   
                      'Title': ''+Title10Char+'',
                      'Enabled': 'True',
                      'SIPTag': ''+SIPTag10Char+'',
                      'SIPEnabled': 'True',
    
                    }
        
        # Url For Update Custom Fields    
        UrlforUpdateCustomFields = '/CustomFields/Update'
        URL = ''+common.Domain+''+UrlforUpdateCustomFields +''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
            
        status = 'Failed'
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 409:
                            print(common.SuccessMessage)
                            status = 'Passed'
                    else:
                     status = 'Failed'
                     assert False
                        
            # Write Output Result in Excel File
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
                TestCasesStatus=False
                
    # Test Case End
    
   
class test_2_GETCustomFields(TestCase):
     
    # Url For Get Custom Fields   
    UrlForGetAllCustomFields = '/CustomFields/Get'
    
    # Start Test Case No 34-01
    def testcase_01_GETCustomFields(self, TestCasesStatus=True):
        
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '34-01'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CustomFields' , 'Using Get Method Get CustomFields Data' , 'Get all Data of CustomFields.')
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
    
                    }
        
        # Url For Get Custom Fields
        URL = ''+common.Domain+''+self.UrlForGetAllCustomFields+''
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    if resp['list'] == '[]':
                        print(common.SuccessMessage)
                        status = 'Passed - But List is Empty'
                    else:
                        print(common.SuccessMessage)
                        status = 'Passed'
                else:
                    status = 'Failed'
                    
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['34-Custom Fields']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Test Case End