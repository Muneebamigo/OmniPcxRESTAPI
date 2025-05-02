'''
Created on Aug 2, 2018
  
update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time, requests
from Settings import CommonFunctions as CF
from openpyxl import load_workbook
from unittest import TestCase
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from test_Provisioning import test_05_TeamConfigurations
from Key import config
SheetName=	'41-Calls Count'

class test_1_Calls_Count(TestCase):

    UrlForGetCallCount = '/Calls/CallsCount'

    def testcase_01_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-01'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes' :"10"

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_02_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-02'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with CallDateSearchCriteria 7.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"7",
                      'StartDate':"18-07-2018 00:00:00",
                      'EndDate':"20-07-2018 00:00:00",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': "10"

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_03_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-03'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with CallDateSearchCriteria 7 and invalid/null date.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"7",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': ""

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_04_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-04'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with CallDateSearchCriteria 6.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"6",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"9",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': "10"

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_05_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-05'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with CallDateSearchCriteria 6 and invalid/null ParamNumberOfDays.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"6",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': "10"

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'

                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_06_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-06'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with CallDurationCriteria 4.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"0",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"0.5-1",
                      'CallDurationCriteria':"4",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_07_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-07'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with CallDurationCriteria 4 and CallDuration is null.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"0",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"4",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False


            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['41-Calls Count']
                first_column = ws['B']
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False

    def testcase_08_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-08'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with PrimaryPBXIPCriteria = 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"3",
                      'PrimaryPBXIP':"172.20.1.214",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_09_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-09'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get Recorded calls with PrimaryPBXIPCriteria = 2 and PrimaryPBXIP null.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"2",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_10_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-10'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'ServerRoleCriteria is 2 and server role 0 .')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"2",
                      'ServerID':"0",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_11_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-11'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'ServerRoleCriteria is 2 and server role null .')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"2",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'

                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_12_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-12'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when invalid/non existing RuleID.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"812",
                      'SIPTrunkDDI':"",
                      'RuleID':'123456',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_13_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-13'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when  EncryptedCallsCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"0",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_14_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-14'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when  EncryptedCallsCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"1",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_15_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-15'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when GroupSearchOption is 0.')
        # Add Teams Function calling
        teamfunction=test_05_TeamConfigurations.Test_1_AddTeamCofiguration()
        TeamName=teamfunction.test_01_AddTeam(common.PrereqTestCasesStatusUpdate)

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':''+TeamName+'',
                      'GroupSearchOption':"0",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'1',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_16_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-16'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when GroupSearchOption is 1.')
        cursor = common.DBConnectivity()
        # Add Teams Function calling
        teamfunction=test_05_TeamConfigurations.Test_1_AddTeamCofiguration()
        TeamName=teamfunction.test_01_AddTeam(common.PrereqTestCasesStatusUpdate)

        SQLCommand = ("Select Groupid from Groups Where Name = '"+TeamName+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        TeamID = str(vals[0])
        cursor.commit()

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':''+TeamName+'',
                      'GroupSearchOption':"1",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':''+TeamID+'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_17_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-17'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when AssociatedAgentCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"0",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_18_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-18'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when AssociatedAgentCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"1",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_19_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-19'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts CallDateSearchCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"1",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_20_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-20'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts CallDateSearchCriteria is 3.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"3",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_21_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-21'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts CallDateSearchCriteria is 5.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"4",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_22_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-22'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when CallTypeSearchCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"1",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_23_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-23'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when CallTypeSearchCriteria is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"2",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_24_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-24'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when DeviceHangUpCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"0",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_25_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-25'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when DeviceHangUpCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"1",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_26_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-26'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when  CallDirectionCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"1",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_27_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-27'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when  CallDirectionCriteria is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"2",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_28_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-28'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when  CallDirectionCriteria is invalid or non existing.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"112",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_29_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-29'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when CallDurationCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"50",
                      'CallDurationCriteria':"0",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_30_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-30'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when CallDurationCriteria is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"50",
                      'CallDurationCriteria':"2",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_31_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-31'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ScreenCaptureCallsCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"0",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_32_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-32'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ScreenCaptureCallsCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"1",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_33_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-33'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ArchivedCallsCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"0",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_34_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-34'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ArchivedCallsCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"1",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_35_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-35'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ArchivedCallsCriteria is invalid or non existing.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2212",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_36_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-36'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when RecordingInterface is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"0",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_37_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-37'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when RecordingInterface is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"1",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_38_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-38'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when RecordingInterface is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"2",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_39_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-39'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when RecordingInterface is 3.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"3",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_40_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-40'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when RecordingInterface is 4.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"4",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_41_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-41'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when RecordingInterface is 5.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"5",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_42_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-42'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ServerRole is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"1",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_43_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-43'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ServerRole is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"2",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_44_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-44'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when DeviceSearchCriteria is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"1212",
                      'DeviceSearchCriteria':"0",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_45_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-45'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when DeviceSearchCriteria is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"1313",
                      'DeviceSearchCriteria':"2",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_46_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-46'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when DeviceSearchCriteria is 3.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"1212",
                      'DeviceSearchCriteria':"3",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': ''



                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_47_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-47'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when SortDirection is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"1",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_48_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-48'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ScoredCallOption is 0.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"0",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_49_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-49'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when ScoredCallOption is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"1",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_50_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-50'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when PrimaryPBXIPCriteria is invalid on non existing.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"1212",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_51_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-51'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when SpeechAnalyticsCriteria is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'1',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_52_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-52'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when SpeechAnalyticsCriteria is 2.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'2',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_53_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-53'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when SpeechAnalyticsCriteria is 3.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'3',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_54_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-54'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts when SpeechAnalyticsCriteria is 4.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'4',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_55_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-55'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts with invalid or non existing SpeechAnalyticsCriteria value.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'123456',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_56_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-56'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts with Null SpeechAnalyticsCriteria value.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"5",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_57_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-57'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts with callstatus is 1.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"1",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_58_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-58'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts with callstatus is 6.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"6",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_59_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-59'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts with callstatus is 15.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"15",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_60_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-60'
        common = CF.CommonFunctions()
        common.Header('Calls Count' , 'Using Get Method to Calls Count' , 'Get all counts with callstatus is 19.')

        Parameters = {'AuthToken':config.sessionkeysite,
                      'AuthUser':config.auth_user,
                      'SiteCode':"",
                      'FirstName':"",
                      'FirstNameCriteria':"8",
                      'LastName':"",
                      'LastNameCriteria':"8",
                      'GroupName':"",
                      'GroupSearchOption':"2",
                      'CalledBy':"",
                      'CalledByCriteria':"8",
                      'CalledTo':"",
                      'CalledToCriteria':"8",
                      'EncryptedCallsCriteria':"2",
                      'AssociatedAgentCriteria':"2",
                      'CallDateSearchCriteria':"5",
                      'StartDate':"",
                      'EndDate':"",
                      'ParamNumberOfDays':"",
                      'CallDuration':"",
                      'CallDurationCriteria':"3",
                      'CallDirectionCriteria':"0",
                      'ArchivedCallsCriteria':"2",
                      'ScreenCaptureCallsCriteria':"2",
                      'DeviceHangUpCriteria':"2",
                      'GlobalCallID':"",
                      'GlobalCallIDSearchCriteria':"8",
                      'CorrelatorID':"",
                      'CorrelatorIDSearchCriteria':"4",
                      'Device':"",
                      'DeviceSearchCriteria':"5",
                      'FlagSearchCriteria':"all",
                      'Notes':"",
                      'NotesSearchCriteria':"8",
                      'CallStatus':"19",
                      'CallTypeSearchCriteria':"0",
                      'RecordingInterface':"6",
                      'SortExpression':"0",
                      'SortDirection':"0",
                      'ScoredCallOption':"2",
                      'PrimaryPBXIPCriteria':"4",
                      'PrimaryPBXIP':"",
                      'SecondaryPBXIPCriteria':"4",
                      'SecondaryPBXIP':"",
                      'ServerRoleCriteria':"0",
                      'ServerRole':"0",
                      'IncludeExternalArchivedCall':"true",
                      'TotalCallShow':"100",
                      'ServerIDCriteria':"8",
                      'ServerID':"",
                      'PageNumber':"1",
                      'CallsPerPage':"50",
                      'CustomField1_Value':"",
                      'CustomField1_Criteria':"8",
                      'CustomField2_Value':"",
                      'CustomField2_Criteria':"8",
                      'CustomField3_Value':"",
                      'CustomField3_Criteria':"8",
                      'CustomField4_Value':"",
                      'CustomField4_Criteria':"8",
                      'CustomField5_Value':"",
                      'CustomField5_Criteria':"8",
                      'CustomField6_Value':"",
                      'CustomField6_Criteria':"8",
                      'CustomField7_Value':"",
                      'CustomField7_Criteria':"8",
                      'CustomField8_Value':"",
                      'CustomField8_Criteria':"8",
                      'CustomField9_Value':"",
                      'CustomField9_Criteria':"8",
                      'CustomField10_Value':"",
                      'CustomField10_Criteria':"8",
                      'SIPTrunkDDI':"",
                      'RuleID':'',
                      'SpeechAnalyticsCriteria':'0',
                      'BoardID':'',
                      'TrunkGroupID':'',
                      'Channel':'',
                      'ProACD':'',
                      'ProACDSearchCriteria':'5',
                      'CallingPartyName':'',
                      'CallingPartyNameSearchCriteria':'8',
                      'GroupID':'',
                      'ParamNumberOfMinutes': '10'

                     }

        URL = ''+common.Domain+''+self.UrlForGetCallCount+''
        response = requests.get(URL, headers=Parameters)

        resp=response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False

    def testcase_61_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-61'
        common = CF.CommonFunctions()
        common.Header('Calls Count', 'Using Get Method to Calls Count', 'Get all counts via server session')

        Parameters = {'AuthToken': config.sessionkey,
                      'AuthUser': config.auth_user,
                      'SiteCode': "",
                      'FirstName': "",
                      'FirstNameCriteria': "8",
                      'LastName': "",
                      'LastNameCriteria': "8",
                      'GroupName': "",
                      'GroupSearchOption': "2",
                      'CalledBy': "",
                      'CalledByCriteria': "8",
                      'CalledTo': "",
                      'CalledToCriteria': "8",
                      'EncryptedCallsCriteria': "2",
                      'AssociatedAgentCriteria': "2",
                      'CallDateSearchCriteria': "8",
                      'StartDate': "",
                      'EndDate': "",
                      'ParamNumberOfDays': "",
                      'CallDuration': "",
                      'CallDurationCriteria': "3",
                      'CallDirectionCriteria': "0",
                      'ArchivedCallsCriteria': "2",
                      'ScreenCaptureCallsCriteria': "2",
                      'DeviceHangUpCriteria': "2",
                      'GlobalCallID': "",
                      'GlobalCallIDSearchCriteria': "8",
                      'CorrelatorID': "",
                      'CorrelatorIDSearchCriteria': "4",
                      'Device': "",
                      'DeviceSearchCriteria': "5",
                      'FlagSearchCriteria': "all",
                      'Notes': "",
                      'NotesSearchCriteria': "8",
                      'CallStatus': "5",
                      'CallTypeSearchCriteria': "0",
                      'RecordingInterface': "6",
                      'SortExpression': "0",
                      'SortDirection': "0",
                      'ScoredCallOption': "2",
                      'PrimaryPBXIPCriteria': "4",
                      'PrimaryPBXIP': "",
                      'SecondaryPBXIPCriteria': "4",
                      'SecondaryPBXIP': "",
                      'ServerRoleCriteria': "0",
                      'ServerRole': "0",
                      'IncludeExternalArchivedCall': "true",
                      'TotalCallShow': "100",
                      'ServerIDCriteria': "8",
                      'ServerID': "",
                      'PageNumber': "1",
                      'CallsPerPage': "50",
                      'CustomField1_Value': "",
                      'CustomField1_Criteria': "8",
                      'CustomField2_Value': "",
                      'CustomField2_Criteria': "8",
                      'CustomField3_Value': "",
                      'CustomField3_Criteria': "8",
                      'CustomField4_Value': "",
                      'CustomField4_Criteria': "8",
                      'CustomField5_Value': "",
                      'CustomField5_Criteria': "8",
                      'CustomField6_Value': "",
                      'CustomField6_Criteria': "8",
                      'CustomField7_Value': "",
                      'CustomField7_Criteria': "8",
                      'CustomField8_Value': "",
                      'CustomField8_Criteria': "8",
                      'CustomField9_Value': "",
                      'CustomField9_Criteria': "8",
                      'CustomField10_Value': "",
                      'CustomField10_Criteria': "8",
                      'SIPTrunkDDI': "",
                      'RuleID': '',
                      'SpeechAnalyticsCriteria': '0',
                      'BoardID': '',
                      'TrunkGroupID': '',
                      'Channel': '',
                      'ProACD': '',
                      'ProACDSearchCriteria': '5',
                      'CallingPartyName': '',
                      'CallingPartyNameSearchCriteria': '8',
                      'GroupID': '',
                      'ParamNumberOfMinutes': "10"

                      }

        URL = '' + common.Domain + '' + self.UrlForGetCallCount + ''
        response = requests.get(URL, headers=Parameters)

        resp = response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 401:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False

    def testcase_62_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-62'
        common = CF.CommonFunctions()
        common.Header('Calls Count', 'Using Get Method to Calls Count', 'Get all counts when CallDateSearchCriteria i 8 and ParamNumberOfMinutes is 10')

        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'SiteCode': "",
                      'FirstName': "",
                      'FirstNameCriteria': "8",
                      'LastName': "",
                      'LastNameCriteria': "8",
                      'GroupName': "",
                      'GroupSearchOption': "2",
                      'CalledBy': "",
                      'CalledByCriteria': "8",
                      'CalledTo': "",
                      'CalledToCriteria': "8",
                      'EncryptedCallsCriteria': "2",
                      'AssociatedAgentCriteria': "2",
                      'CallDateSearchCriteria': "8",
                      'StartDate': "",
                      'EndDate': "",
                      'ParamNumberOfDays': "",
                      'CallDuration': "",
                      'CallDurationCriteria': "3",
                      'CallDirectionCriteria': "0",
                      'ArchivedCallsCriteria': "2",
                      'ScreenCaptureCallsCriteria': "2",
                      'DeviceHangUpCriteria': "2",
                      'GlobalCallID': "",
                      'GlobalCallIDSearchCriteria': "8",
                      'CorrelatorID': "",
                      'CorrelatorIDSearchCriteria': "4",
                      'Device': "",
                      'DeviceSearchCriteria': "5",
                      'FlagSearchCriteria': "all",
                      'Notes': "",
                      'NotesSearchCriteria': "8",
                      'CallStatus': "5",
                      'CallTypeSearchCriteria': "0",
                      'RecordingInterface': "6",
                      'SortExpression': "0",
                      'SortDirection': "0",
                      'ScoredCallOption': "2",
                      'PrimaryPBXIPCriteria': "4",
                      'PrimaryPBXIP': "",
                      'SecondaryPBXIPCriteria': "4",
                      'SecondaryPBXIP': "",
                      'ServerRoleCriteria': "0",
                      'ServerRole': "0",
                      'IncludeExternalArchivedCall': "true",
                      'TotalCallShow': "100",
                      'ServerIDCriteria': "8",
                      'ServerID': "",
                      'PageNumber': "1",
                      'CallsPerPage': "50",
                      'CustomField1_Value': "",
                      'CustomField1_Criteria': "8",
                      'CustomField2_Value': "",
                      'CustomField2_Criteria': "8",
                      'CustomField3_Value': "",
                      'CustomField3_Criteria': "8",
                      'CustomField4_Value': "",
                      'CustomField4_Criteria': "8",
                      'CustomField5_Value': "",
                      'CustomField5_Criteria': "8",
                      'CustomField6_Value': "",
                      'CustomField6_Criteria': "8",
                      'CustomField7_Value': "",
                      'CustomField7_Criteria': "8",
                      'CustomField8_Value': "",
                      'CustomField8_Criteria': "8",
                      'CustomField9_Value': "",
                      'CustomField9_Criteria': "8",
                      'CustomField10_Value': "",
                      'CustomField10_Criteria': "8",
                      'SIPTrunkDDI': "",
                      'RuleID': '',
                      'SpeechAnalyticsCriteria': '0',
                      'BoardID': '',
                      'TrunkGroupID': '',
                      'Channel': '',
                      'ProACD': '',
                      'ProACDSearchCriteria': '5',
                      'CallingPartyName': '',
                      'CallingPartyNameSearchCriteria': '8',
                      'GroupID': '',
                      'ParamNumberOfMinutes': "10"

                      }

        URL = '' + common.Domain + '' + self.UrlForGetCallCount + ''
        response = requests.get(URL, headers=Parameters)

        resp = response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False

    def testcase_63_Calls_Count(self, TestCasesStatus=True):

        starttime = time.process_time()
        TestCaseID = '41-63'
        common = CF.CommonFunctions()
        common.Header('Calls Count', 'Using Get Method to Calls Count', 'Get all counts when CallDateSearchCriteria i 8 and ParamNumberOfMinutes is empty')

        Parameters = {'AuthToken': config.sessionkeysite,
                      'AuthUser': config.auth_user,
                      'SiteCode': "",
                      'FirstName': "",
                      'FirstNameCriteria': "8",
                      'LastName': "",
                      'LastNameCriteria': "8",
                      'GroupName': "",
                      'GroupSearchOption': "2",
                      'CalledBy': "",
                      'CalledByCriteria': "8",
                      'CalledTo': "",
                      'CalledToCriteria': "8",
                      'EncryptedCallsCriteria': "2",
                      'AssociatedAgentCriteria': "2",
                      'CallDateSearchCriteria': "8",
                      'StartDate': "",
                      'EndDate': "",
                      'ParamNumberOfDays': "",
                      'CallDuration': "",
                      'CallDurationCriteria': "3",
                      'CallDirectionCriteria': "0",
                      'ArchivedCallsCriteria': "2",
                      'ScreenCaptureCallsCriteria': "2",
                      'DeviceHangUpCriteria': "2",
                      'GlobalCallID': "",
                      'GlobalCallIDSearchCriteria': "8",
                      'CorrelatorID': "",
                      'CorrelatorIDSearchCriteria': "4",
                      'Device': "",
                      'DeviceSearchCriteria': "5",
                      'FlagSearchCriteria': "all",
                      'Notes': "",
                      'NotesSearchCriteria': "8",
                      'CallStatus': "5",
                      'CallTypeSearchCriteria': "0",
                      'RecordingInterface': "6",
                      'SortExpression': "0",
                      'SortDirection': "0",
                      'ScoredCallOption': "2",
                      'PrimaryPBXIPCriteria': "4",
                      'PrimaryPBXIP': "",
                      'SecondaryPBXIPCriteria': "4",
                      'SecondaryPBXIP': "",
                      'ServerRoleCriteria': "0",
                      'ServerRole': "0",
                      'IncludeExternalArchivedCall': "true",
                      'TotalCallShow': "100",
                      'ServerIDCriteria': "8",
                      'ServerID': "",
                      'PageNumber': "1",
                      'CallsPerPage': "50",
                      'CustomField1_Value': "",
                      'CustomField1_Criteria': "8",
                      'CustomField2_Value': "",
                      'CustomField2_Criteria': "8",
                      'CustomField3_Value': "",
                      'CustomField3_Criteria': "8",
                      'CustomField4_Value': "",
                      'CustomField4_Criteria': "8",
                      'CustomField5_Value': "",
                      'CustomField5_Criteria': "8",
                      'CustomField6_Value': "",
                      'CustomField6_Criteria': "8",
                      'CustomField7_Value': "",
                      'CustomField7_Criteria': "8",
                      'CustomField8_Value': "",
                      'CustomField8_Criteria': "8",
                      'CustomField9_Value': "",
                      'CustomField9_Criteria': "8",
                      'CustomField10_Value': "",
                      'CustomField10_Criteria': "8",
                      'SIPTrunkDDI': "",
                      'RuleID': '',
                      'SpeechAnalyticsCriteria': '0',
                      'BoardID': '',
                      'TrunkGroupID': '',
                      'Channel': '',
                      'ProACD': '',
                      'ProACDSearchCriteria': '5',
                      'CallingPartyName': '',
                      'CallingPartyNameSearchCriteria': '8',
                      'GroupID': '',
                      'ParamNumberOfMinutes': ""

                      }

        URL = '' + common.Domain + '' + self.UrlForGetCallCount + ''
        response = requests.get(URL, headers=Parameters)

        resp = response.json()
        showcode = str(resp['ResponseCode'])

        if TestCasesStatus == True:
            try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                    status = 'Failed'
                    assert False

            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus = False