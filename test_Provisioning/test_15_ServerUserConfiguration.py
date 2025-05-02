'''
Created on Jul 20, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time,requests
from Settings import CommonFunctions as CF
from unittest import TestCase
from openpyxl import load_workbook
from test_Provisioning import test_14_ServerPermission as serverpermissionfunctions
from test_Provisioning import test_01_SystemSettings
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from Key import config
SheetName=	'15-Server User Configuration'

class Test_1_AddServerUser(TestCase):
    
    # Start Test Case No 15-01
    def test_01_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        
        TestCaseID = '15-01'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Default Admin User')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
      
        else:
            TestCasesStatus=False 
            
        return FirstName, Username
        # Test Case End
  
    # Start Test Case No 15-04
    def test_02_AddServerUser_With_Default_Admin_Duplicate(self):
       
        TestCaseID = '15-04'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Duplicate Server User with Default Admin User')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        FirstName, Username= Test_1_AddServerUser.test_01_AddServerUser_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        
        # Generate Simple Character String Limit 10 Characters
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
            if resp['ResponseCode'] == 409:
                print(common.SuccessMessage)
                status ='Passed'
                        
            else:
                status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End
                        
    # Start Test Case No 15-14
    def test_03_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-14'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Null/Empty FirstName')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = ''
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
            
    # Start Test Case No 15-15
    def test_04_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-15'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with NULL/Empty LastName')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = ''
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
            
    # Start Test Case No 15-16
    def test_05_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-16'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Null/Emprt UserName')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = ''
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
            
    # Start Test Case No 15-17
    def test_06_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-17'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Less then 7 character password')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = 'abc'
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
                        
    # Start Test Case No 15-18
    def test_07_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-18'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with invalid email format')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = 'abc123'
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
            
    # Start Test Case No 15-02
    def test_08_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Default Admin User when server role as secondary configured.')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        
        SSF=test_01_SystemSettings.test_1_UpdateSystemSettings()
        SSF.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
            
    # Start Test Case No 15-03
    def test_09_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        TestCaseID = '15-03'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Default Admin User when server role as branch configured.')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0]) 
        
        SSF=test_01_SystemSettings.test_1_UpdateSystemSettings()
        SSF.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 403:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
  
        else:
            TestCasesStatus=False
            
            
    # Start Test Case No 15-19
    def test_10_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        
        TestCaseID = '15-19'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Default Admin User when default landing page is Tenants.')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': '0'
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 200:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
      
        else:
            TestCasesStatus=False 
            
        return FirstName, Username
        # Test Case End
        
        
    # Start Test Case No 15-20
    def test_11_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        
        TestCaseID = '15-20'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Default Admin User when invalid or non existing default landing page value.')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': '17'
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
      
        else:
            TestCasesStatus=False
            
        # Test Case End
        
        
    # Start Test Case No 15-21
    def test_12_AddServerUser_With_Default_Admin(self, TestCasesStatus=True):
        
        
        TestCaseID = '15-21'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Adding Method of Server User Configuration', 'Add Server User with Default Admin User when default landing page value is invalid format as a string.')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        serverpermissionname = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+serverpermissionname+"';")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Username = common.GenrateSimpleStringLimit10()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Username': ''+Username+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': 'abcdef'
        
                     }
        #URL for Add Server User
        UrlAddServerUser ='/ServerUser/Add'
        URL = ''+common.Domain+''+UrlAddServerUser+''
        # Hit API Through Methods
        response = requests.post(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                    if resp['ResponseCode'] == 400:
                        print(common.SuccessMessage)
                        status ='Passed'
                            
                    else:
                     status = 'Failed'
                     assert False
            # Write Output Result in Excel File            
            finally:
                    common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
      
        else:
            TestCasesStatus=False
            
        # Test Case End

class Test_2_GetServerUser(TestCase):
    
    #URL for Get Server User
    UrlGetServerUser='/ServerUser/Get'
    UrlGetServerUserById='/ServerUser/Get/'
    UrlGetServerUserByName='/ServerUser/Get/?Username='
    # Start Test Case No 15-05
    def test_01_GetAllServerUser_With_Default_Admin(self):
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '15-05'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Get Method of Server User Configuration', 'Getting all list of Server User with Default Admin User')
     
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'id':'',
                      'username': ''
        
                     }
        
        URL = ''+common.Domain+''+self.UrlGetServerUser+''
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End

    # Start Test Case No 15-06
    def test_02_GetServerUserById_With_Default_Admin(self):
        
        TestCaseID = '15-06'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Get Method of Server User Configuration', 'Getting Server User By Id with Default Admin User')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        
        FirstName, Username = Test_1_AddServerUser.test_01_AddServerUser_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        Username=Username
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("Select UserId from Users Where FirstName = '"+FirstName+"'; ")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        UserId = str(vals[0])
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'id':''+UserId+'',
                      'username': ''
        
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlGetServerUserById+UserId
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File           
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End
                        
    # Start Test Case No 15-07
    def test_03_GetServerUserByName_With_Default_Admin(self):
       
        TestCaseID = '15-07'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Get Method of Server User Configuration', 'Getting Server User By username with Default Admin User')
        
        FirstName, Username = Test_1_AddServerUser.test_01_AddServerUser_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        FirstName=FirstName
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': '',
                      'username': ''+Username+''
                      
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlGetServerUserByName+"'"+Username+"'"
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End

    # Start Test Case No 15-08
    def test_04_GetServerUserById_With_Default_Admin(self):
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '15-08'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Get Method of Server User Configuration', 'Getting Server User By invalid Id with Default Admin User')
        
        UserId = '123456'
    
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'id':''+UserId+'',
                      'username': ''
        
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlGetServerUserById+UserId
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File           
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End

    # Start Test Case No 15-09
    def test_05_GetServerUserByName_With_Default_Admin(self):
        # Test Case Start Time
        starttime = time.process_time()
        TestCaseID = '15-09'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Get Method of Server User Configuration', 'Getting Server User By invalid username with Default Admin User')
        
        Username = 'abc123abc123'
    
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'Id': '',
                      'username': ''+Username+''
                                          
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlGetServerUserByName+"'"+Username+"'"
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End
                        
class Test_3_UpdateServerUser(TestCase):
    #URL for Update Server User
    UrlUpdateServerUser ='/ServerUser/Update'
    # Start Test Case No 15-10
    def test_01_UpadteServerUser_With_Default_Admin(self):
        
        TestCaseID = '15-10'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Update Method of Server User Configuration', 'Update Server User with Default Admin User')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        FirstName, Username = Test_1_AddServerUser.test_01_AddServerUser_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        Username=Username
        SQLCommand = ("Select UserId from Users where FirstName = '"+FirstName+"'; ")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        UserId = str(vals[0])
        # Server Permission Function calling
        ServerPermissionFunctions = serverpermissionfunctions.Test_1_AddServerPermission()
        Name = ServerPermissionFunctions.test_01_AddPermission_With_Default_Admin(common.PrereqTestCasesStatusUpdate)
        # SQL Queries for Data Verification        
        SQLCommand1 = ("Select SG_ID from OPR_Security_Group Where SG_Name = '"+Name+"'")
        cursor.execute(SQLCommand1)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID':''+UserId+'',
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ActiveDirectoryUserEnabled': 'False',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlUpdateServerUser+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End
                        
    # Start Test Case No 15-11
    def test_02_UpadteServerUser_With_Default_Admin(self):
        
        TestCaseID = '15-11'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Update Method of Server User Configuration', 'Update Server User with non existing id')
        
        UserId = '12345'
        
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # SQL Queries for Data Verification        
        SQLCommand1 = ("Select SG_ID from OPR_Security_Group")
        cursor.execute(SQLCommand1)
        vals = cursor.fetchone()
        SG_ID = str(vals[0])
        # Generate Simple Character String Limit 10 Characters
        FirstName = common.GenrateSimpleStringLimit10()
        LastName = common.GenrateSimpleStringLimit10()
        Email = common.GenerateEmail()
        Password = common.GenrateValidPasswordString()
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'ID':''+UserId+'',
                      'FirstName': ''+FirstName+'',
                      'LastName': ''+LastName+'',
                      'Email': ''+Email+'',
                      'Password': ''+Password+'',
                      'ActiveUser': 'True',
                      'PasswordNeverExpireEnabled': 'True',
                      'LoginPasswordChangeEnabled': 'True',
                      'ServerPermission': ''+SG_ID+'',
                      'ADSID': '',
                      'DefaultLandingPage': ''
        
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlUpdateServerUser+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End

class Test_4_DeleteServerUser(TestCase):
    #URL for Delete Server User
    UrlDeleteServerUser='/ServerUser/Delete/'
    
    # Start Test Case No 15-12
    def test_01_DeleteServerUserById_With_Default_Admin(self):
       
        TestCaseID = '15-12'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Delete Method of Server User Configuration', 'Deleting Server User By Id with Default Admin User')
        # Config DB Connectivity Function calling
        cursor = common.StringDBConnectivity()
        # Selecting Random Exisitng security group row from below SQL Query
        SQLCommand = ("SELECT UserId FROM Users WHERE UserId != 1 ORDER BY RANDOM() LIMIT 1;")
        cursor.execute(SQLCommand)
        vals = cursor.fetchone()
        UserId = str(vals[0])
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'id':''+UserId+''
                      
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlDeleteServerUser+UserId
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End
                        
    # Start Test Case No 15-13
    def test_02_DeleteServerUserById_With_Default_Admin(self):
        
        TestCaseID = '15-13'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('Server Users Configuration', 'Calling Delete Method of Server User Configuration', 'Deleting Server User By non existing/invalid Id')
        
        UserId = '12345'
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkey,
                      'AuthUser':config.auth_user,
                      'id':''+UserId+''
                      
                     }
        #URL
        URL = ''+common.Domain+''+self.UrlDeleteServerUser+UserId
        # Hit API Through Methods
        response = requests.delete(URL, headers=Parameters)
        # API Response in JSon Format
        resp = response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        try:
                if resp['ResponseCode'] == 400:
                    print(common.SuccessMessage)
                    status ='Passed'
                        
                else:
                    status='Failed'
        # Write Output Result in Excel File            
        finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['15-Server User Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                        if (first_column[x].value)  == TestCaseID:
                            ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                            ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                            ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                            ProcessingTime = float(str((time.process_time() - starttime + 2)))
                            ws.cell(row=x+1 , column=11).value = ProcessingTime
                            ws.cell(row=x+1 , column=13).value = common.SystemUser
                            ws.cell(row=x+1 , column=14).value = common.WindowServer
                            if(status =='Passed'):
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                            else:
                                ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            ws.cell(row=x+1 , column=19).value = status
                            ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                            ws.cell(row=x+1 , column=18).value = showcode
                            ws.cell(row=x+1 , column=16).value = str(resp)
                            wb.save(''+common.OutPutFilePath+'')
                        # Test Case End          