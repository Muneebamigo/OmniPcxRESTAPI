'''
Created on Jul 24, 2018

update on Feb 17, 2024
this last upadate successfully on 2.5.0.15 release
@author: Muneeb.ahmed
'''

import time, requests, random
from Settings import CommonFunctions as CF
from openpyxl import load_workbook
from unittest import TestCase
from test_Provisioning import test_01_SystemSettings as systemfuncitons
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from Key import config

SheetName=	'33-Call Flags Configuration'

class test_1_GetCallFlagsCongiguration(TestCase):
   
    # Start Test Case No 33-01   
    def testcase_01_GetCallFlagsCongiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-01'
        # Calling Common Functions
        common = CF.CommonFunctions()
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user
                    }
        
        # Url for Get Devices
        UrlForGettingDevice = '/CallFlag/Get/'
        
        URL = ''+common.Domain+''+UrlForGettingDevice
        # Hit API Through Methods
        response = requests.get(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        status = 'Failed'
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                common.UpdateExcelTestCase(SheetName, TestCaseID, URL, Parameters, status, starttime, resp)
        else:
            TestCasesStatus=False
            
    # Test Case End
            
class test_2_UpdateCallFlagsConfiguration(TestCase):
    
    UrlForUpdateCallFlagsConfiguration = '/CallFlag/Update/'
    
    # Start Test Case No 33-02
    def testcase_01_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-02'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System with Valid Input Flags ID.')
        
        # System Settings Function calling        
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        FlagID=str(random.randint(1 , 7))
        FlagText=common.GenrateSimpleStringLimit10()
        SiteCode=''
        
        # Test Case Start Time
        starttime = time.process_time()
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': ''+FlagText+'',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 200:
                            print(common.SuccessMessage)
                            status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Test Case End
            
    # Start Test Case No 33-03
    def testcase_02_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-03'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System with inValid Input Flags ID.')
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)

        FlagID='12345'
        FlagText=common.GenrateSimpleStringLimit10()
        SiteCode=''
        
        # Test Case Start Time
        starttime = time.process_time()
        
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': ''+FlagText+'',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Test Case End
            
    # Start Test Case No 33-04
    def testcase_03_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-04'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System flagtext Characters more then 30.')
        
        # System Settings Function calling
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        
        FlagID=str(random.randint(1 , 7))
        FlagText=common.GenrateDesc250()
        SiteCode=''
        # Test Case Start Time
        starttime = time.process_time()
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': ''+FlagText+'',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        response = requests.put(URL, headers=Parameters)
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Test Case End
    
    # Start Test Case No 33-05
    def testcase_05_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-05'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System when server role as branch recorder.')
        
        # System Settings Function calling        
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_09_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        FlagID=str(random.randint(1 , 7))
        FlagText=common.GenrateSimpleStringLimit10()
        SiteCode=''
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': ''+FlagText+'',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 403:
                            print(common.SuccessMessage)
                            status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Start Test Case No 33-06
    def testcase_06_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-06'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System with InValid Site Code.')
        
        # System Settings Function calling        
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        FlagID=str(random.randint(1 , 7))
        FlagText=common.GenrateSimpleStringLimit10()
        SiteCode='1234567'
        
        # Test Case Start Time
        starttime = time.process_time()
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': ''+FlagText+'',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 500:
                    print(common.SuccessMessage)
                    status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Start Test Case No 33-07
    def testcase_07_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-07'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System when server role as secondary configured.')
        
        # System Settings Function calling        
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_05_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        FlagID=str(random.randint(1 , 7))
        FlagText=common.GenrateSimpleStringLimit10()
        SiteCode=''
        
        # Test Case Start Time
        starttime = time.process_time()
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': ''+FlagText+'',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 403:
                            print(common.SuccessMessage)
                            status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
            
    # Start Test Case No 33-08
    def testcase_08_UpdateCallFlagsConfiguration(self, TestCasesStatus=True):
        
        TestCaseID = '33-08'
        # Calling Common Functions
        common = CF.CommonFunctions()
        common.Header('CallFlags Configuration' , 'Using Put Method Update Call Flags Configuration' , 'Update Call Flags to System with Valid Input Flags ID and in name used double quotes.')
        
        # System Settings Function calling        
        SystemSettingsFunctiopns=systemfuncitons.test_1_UpdateSystemSettings()
        SystemSettingsFunctiopns.testcase_01_UpdateSystemSettings(common.PrereqTestCasesStatusUpdate)
        
        FlagID=str(random.randint(1 , 7))
        SiteCode=''
        
        # Test Case Start Time
        starttime = time.process_time()
        
        # Header Parameters of Rest API
        Parameters = {'AuthToken':config.sessionkeysite,
                    'AuthUser':config.auth_user,
                    'Id': ''+FlagID+'',
                    'SiteCode': ''+SiteCode+'',
                    'FlagText': 'test"name"test',
        
                    }
        
        # Url for Update Devices
        URL = ''+common.Domain+''+self.UrlForUpdateCallFlagsConfiguration+''
        # Hit API Through Methods
        response = requests.put(URL, headers=Parameters)
        # API Response in JSon Format
        resp=response.json()
        showcode = str(resp['ResponseCode'])
        
        # Response Code Verification
        if TestCasesStatus==True:
            try:
                if resp['ResponseCode'] == 400:
                            print(common.SuccessMessage)
                            status = 'Passed'
                else:
                     status = 'Failed'
                     assert False
            
            # Write Output Result in Excel File
            finally:
                wb=load_workbook(''+common.OutPutFilePath+'')
                wb.sheetnames
                ws = wb['33-Call Flags Configuration']
                first_column = ws['B']
                del Parameters["AuthToken"]
                for x in range(len(first_column)):
                    if (first_column[x].value)  == TestCaseID:
                        ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=7).value = str(Parameters).replace(',','\n')
                        ws.cell(row=x+1 , column=8).value = common.ExecutionDate
                        ws.cell(row=x+1 , column=9).value = common.ExecutionTime
                        ProcessingTime = float(str((time.process_time() - starttime + 2)))
                        ws.cell(row=x+1 , column=11).value = ProcessingTime
                        ws.cell(row=x+1 , column=13).value = common.SystemUser
                        ws.cell(row=x+1 , column=14).value = common.WindowServer
                        if(status =='Passed'):
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                        else:
                            ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        ws.cell(row=x+1 , column=19).value = status
                        ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                        ws.cell(row=x+1 , column=18).value = showcode
                        ws.cell(row=x+1 , column=16).value = str(resp)
                        wb.save(''+common.OutPutFilePath+'')
        else:
            TestCasesStatus=False
            
    # Test Case End